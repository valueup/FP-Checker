# -*- coding: utf-8 -*-
"""
comparator.py  v1.0
단계별 FP 산정 결과물 대조 엔진

한쪽 단계의 FP 산정서 여러 개와 다른 단계의 FP 산정서 여러 개를 읽어
단위프로세스를 짝지어 주고, 짝이 없는 것과 값이 달라진 것을 찾아낸다.

  A 그룹 : 예) 설계단계 FP 산정서 3개
  B 그룹 : 예) 종료단계 FP 산정서 5개

짝짓기 결과는 다섯 가지다.
  일치     같은 기능이 양쪽에 있고 값도 같다
  변경     같은 기능이 양쪽에 있으나 FP유형·복잡도·가중치 등이 다르다
  유사     이름이 완전히 같지는 않으나 비슷해서 같은 기능으로 본 것 (사람이 확인해야 함)
  A만 있음 뒤 단계에서 빠졌다 (누락 의심)
  B만 있음 앞 단계에 없던 것이 생겼다 (추가)

양식 사양과 FP 계산(FPR/FPV)은 calculator_ui.py 의 것을 그대로 쓴다.

단독 실행:
  python comparator.py A1.xlsm A2.xlsm -- B1.xlsm B2.xlsm
"""

import os
import re
import sys
import time
import difflib

import openpyxl

# ── 폴더 구조 대응 ────────────────────────────────────────────
# FP-Checker/Calculator, FP-Checker/Comparator 처럼 나뉘어 있어도 서로 불러온다.
# 프로젝트 루트 자체는 넣지 않는다. 루트에 예전 사본이 남아 있으면
# 같은 이름이 먼저 읽혀 옛 코드가 도는 일이 생기기 때문이다.
_HERE = os.path.dirname(os.path.abspath(__file__))
_PKGS = ("Calculator", "Comparator", "DocParser")
for _p in [os.path.join(os.path.dirname(_HERE), _n) for _n in _PKGS] + [_HERE]:
    if os.path.isdir(_p):
        if _p in sys.path:
            sys.path.remove(_p)
        sys.path.insert(0, _p)


def project_root():
    """설정 파일을 둘 위치. 하위 폴더로 나뉘어 있으면 그 위(프로젝트 루트)."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(_HERE) if os.path.basename(_HERE) in _PKGS else _HERE


import calculator_ui as U

VERSION = "1.4"
BLANK_STOP = 2000      # 데이터가 끊긴 뒤 이만큼 빈 행이 이어지면 멈춘다

# 짝짓기 기준
KEY_FIELDS = {
    "proc": (["proc"], "단위프로세스명"),
    "app_proc": (["app", "proc"], "애플리케이션 + 단위프로세스명"),
    "app_biz_proc": (["app", "biz", "proc"], "애플리케이션 + 세부업무 + 단위프로세스명"),
    "proc_type": (["proc", "type"], "단위프로세스명 + FP유형"),
    # 단위프로세스명이 화면 묶음 이름이고 실제 기능은 설명에 적힌 산출물이 많다.
    # (예: 프로세스명 "시설물 마스터 관리" / 설명 "선로시설물 등록(수정) UI-F01-01-02U")
    "proc_desc": (["proc", "desc"], "단위프로세스명 + 설명"),
    "app_proc_desc": (["app", "proc", "desc"], "애플리케이션 + 단위프로세스명 + 설명"),
    "app_biz_proc_desc": (["app", "biz", "proc", "desc"],
                          "애플리케이션 + 세부업무 + 단위프로세스명 + 설명"),
}

# 중복 판정 기준. auto = 짝짓기 기준에 설명과 FP유형을 더해서 본다.
DUP_KEYS = {"auto": "짝짓기 기준 + 설명 + FP유형",
            "same": "짝짓기 기준과 동일"}

# 머리글 표와 맞춤 함수는 calculator_ui 의 것을 그대로 쓴다(두 벌 관리하지 않기 위함).
# 불러오는 시점이 아니라 쓰는 시점에 가져온다. 모듈이 서로를 부르는 차례에 따라
# 아직 다 만들어지지 않은 모듈의 속성을 건드리면 import 가 깨지기 때문이다.


def header_map():
    return U.HEADER_MAP


def map_header(rowvals):
    return U.map_header(rowvals)


# 값 비교 항목
DIFF_FIELDS = [("type", "FP유형"), ("cx", "복잡도"), ("wt", "가중치"),
               ("ftr", "RET/FTR"), ("det", "DET"), ("dev", "개발유형"),
               ("app", "애플리케이션명"), ("biz", "세부업무명")]

DEFAULT_OPTS = {"key": "app_proc", "dup": "auto", "threshold": 0.85, "fuzzy": True,
                "ignore_space": True, "ignore_bracket": True,
                "ignore_case": True,
                "diff": ["type", "cx", "wt"]}


# ---------------------------------------------------------------- 읽기

def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _weight(method, ttype, ftr, det):
    """양식의 산정방법에 맞춰 복잡도와 가중치를 구한다."""
    t = (ttype or "").upper()
    if not t:
        return "", None
    if method == "간이법":
        return "", U.SIMPLE_WEIGHT.get(t)
    f, d = _num(ftr), _num(det)
    if f is None or d is None:
        return "", None
    g = U.fp_rating(t, f, d)
    if g is None or g == 0:
        return ("0" if g == 0 else ""), None
    return g, U.fp_value(t, g)


def _cell(vals, col):
    return vals[col - 1] if len(vals) >= col else None


def load_rows(path, form_key=None):
    """산정양식 한 개를 읽어 기능 목록을 돌려준다.

    읽기전용 모드로 한 번만 훑는다. 같은 파일을 보통 모드로 열면 15배쯤 느리다.
    산정방법(간이법/상세법) 셀은 양식 사양의 자리 힌트로 바로 읽고,
    힌트가 없을 때만 유효성검사를 뒤지는 느린 경로로 내려간다.
    """
    bad = U.check_path(path)
    if bad:
        raise ValueError(bad.replace("\n", " "))
    form_key = form_key or U.guess_form(path)
    if form_key not in U.FORMS or not U.is_ready(form_key):
        raise ValueError("다룰 수 없는 양식입니다: %s" % (form_key or "판단 불가"))
    spec = U.FORMS[form_key]["fp"]
    spec_cx = {c["k"]: c["x"] for c in spec["cols"]}
    maxc = max(max(spec_cx.values()), 30)      # 열이 밀려 있을 수 있으므로 넉넉히 본다
    in_keys = ("app", "biz", "proc", "desc", "dev", "chg", "type", "ftr", "det", "remark")

    mcell = spec.get("method_cell")
    mrow = mcol = None
    if mcell:
        m = re.match(r"^([A-Z]+)(\d+)$", mcell)
        if m:
            mcol = openpyxl.utils.column_index_from_string(m.group(1))
            mrow = int(m.group(2))

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = None
    for n in spec["sheets"]:
        if n in wb.sheetnames:
            ws = wb[n]
            break
    if ws is None:
        wb.close()
        raise ValueError("FP 시트를 찾지 못했습니다: %s" % os.path.basename(path))

    METHODS = ("간이법", "상세법")
    method = spec.get("method", "상세법")
    method_found = not spec.get("method_dv")
    method_src = "양식 기본값"
    hrow = None
    blank = 0
    rows = []
    name = os.path.basename(path)
    wide = max(maxc, mcol or 0)

    cx = dict(spec_cx)
    mapped = False
    for r, vals in enumerate(ws.iter_rows(min_col=1, max_col=wide, values_only=True), start=1):
        if hrow is None:
            if not method_found:
                # 자리 힌트로 먼저 보고, 열이 밀렸을 수 있으니 그 줄에서 값으로도 찾는다
                v = U._s(_cell(vals, mcol)) if (mrow and r == mrow and mcol) else ""
                if v not in METHODS:
                    v = ""
                    for x in vals:
                        if U._s(x) in METHODS:
                            v = U._s(x)
                            break
                if v:
                    method = v
                    method_found = True
                    method_src = "파일의 산정방법 칸"
            got = map_header(vals)          # 열 위치가 아니라 머리글 글자로 찾는다
            if got:
                hrow = r
                cx = got
                mapped = True
            continue
        rec = {k: U._s(_cell(vals, cx[k])) for k in in_keys if k in cx}
        if not any(rec.get(k) for k in ("app", "biz", "proc", "desc", "type")):
            blank += 1
            if blank > BLANK_STOP and rows:
                break          # 서식만 남아 시트 끝이 부풀려진 파일 대비
            continue
        rec["type"] = rec.get("type", "").upper()
        rec["src"] = name
        rec["path"] = path
        rec["no"] = r
        rec["form"] = form_key
        rows.append(rec)
        blank = 0
    ws_title = ws.title
    wb.close()

    if hrow is None:
        raise ValueError("표 머리글을 찾지 못했습니다: %s" % name)

    if not method_found:                      # 힌트가 없는 양식만 느린 경로
        wb2 = openpyxl.load_workbook(path, data_only=True)
        ws2 = wb2[ws_title]
        mc = U.find_method_cell(ws2, spec["method_dv"])
        if mc:
            method = U._s(ws2[mc].value) or method
        wb2.close()

    if spec.get("method_dv") and not method_found:
        has_ftr = any(rec.get("ftr") for rec in rows)
        has_det = any(rec.get("det") for rec in rows)
        method = "상세법" if (has_ftr and has_det) else "간이법"
        method_src = "RET/FTR·DET 유무로 추정"

    for rec in rows:
        g, w = _weight(method, rec["type"], rec.get("ftr"), rec.get("det"))
        rec["cx"] = g
        rec["wt"] = w

    return {"path": path, "name": name, "form": form_key,
            "formName": U.FORMS[form_key].get("name", form_key),
            "sheet": ws_title, "headerRow": hrow, "method": method,
            "methodSrc": method_src, "cols": cx, "mapped": mapped, "rows": rows}


def peek(path, form_key=None):
    """파일을 열어 무엇으로 읽히는지만 알려 준다(화면 표시용)."""
    d = load_rows(path, form_key)
    have = [k for k in ("dev", "chg", "ftr", "det", "cx", "wt", "wtadj") if k in d["cols"]]
    return {"form": d["form"], "formName": d["formName"], "sheet": d["sheet"],
            "method": d["method"], "methodSrc": d["methodSrc"],
            "headerRow": d["headerRow"], "count": len(d["rows"]),
            "fp": round(sum(r["wt"] or 0 for r in d["rows"]), 1), "cols": have}


class Cancelled(Exception):
    pass


def _tick(progress, stage, done, total, detail=""):
    if progress:
        progress(stage, done, total, detail)


def load_group(files, label, progress=None, cancel=None, on_file=None):
    """한 단계(그룹)의 파일 여러 개를 읽어 합친다."""
    out = {"label": label, "files": [], "rows": [], "errors": []}
    total = len(files)
    for i, f in enumerate(files):
        if cancel and cancel():
            raise Cancelled()
        _tick(progress, "%s 읽는 중" % label, i, total,
              os.path.basename(f.get("path") if isinstance(f, dict) else f))
        path = f.get("path") if isinstance(f, dict) else f
        form = f.get("form") if isinstance(f, dict) else None
        try:
            d = load_rows(path, form)
        except Exception as e:
            out["errors"].append({"path": path, "msg": str(e)})
            continue
        out["files"].append({"path": path, "name": d["name"], "form": d["form"],
                             "formName": d["formName"], "sheet": d["sheet"],
                             "method": d["method"], "methodSrc": d.get("methodSrc", ""),
                             "count": len(d["rows"])})
        out["rows"].extend(d["rows"])
        if on_file:
            on_file(out)
    _tick(progress, "%s 읽는 중" % label, total, total, "")
    return out


# ---------------------------------------------------------------- 짝짓기

_BRACKET = re.compile(r"[\(\)\[\]\{\}（）［］【】<>《》]")


def norm(s, opts):
    t = U._s(s)
    if opts.get("ignore_bracket", True):
        t = _BRACKET.sub(" ", t)
    if opts.get("ignore_space", True):
        t = re.sub(r"[\s\u3000_\-·:/]+", "", t)
    else:
        t = re.sub(r"\s+", " ", t).strip()
    if opts.get("ignore_case", True):
        t = t.upper()
    return t


def row_key(row, opts, fields=None):
    if fields is None:
        fields = KEY_FIELDS.get(opts.get("key", "app_proc"), KEY_FIELDS["app_proc"])[0]
    return "|".join(norm(row.get(f), opts) for f in fields)


def dup_fields(opts):
    """중복 판정에 쓸 항목. 짝짓기 기준과 따로 둔다."""
    d = opts.get("dup", "auto")
    key = opts.get("key", "app_proc")
    if d in KEY_FIELDS:
        return KEY_FIELDS[d][0], KEY_FIELDS[d][1]
    if d == "same":
        return KEY_FIELDS[key][0], "짝짓기 기준과 동일 (%s)" % KEY_FIELDS[key][1]
    f = list(KEY_FIELDS.get(key, KEY_FIELDS["app_proc"])[0])
    for extra in ("desc", "type"):
        if extra not in f:
            f.append(extra)
    return f, "짝짓기 기준 + 단위프로세스 설명 + FP유형"


def row_text(row, opts):
    """유사도 비교에 쓸 글자열."""
    return norm(" ".join(U._s(row.get(f)) for f in ("app", "biz", "proc")), opts)


def proc_text(row, opts):
    """단위프로세스명에서 앞에 붙은 애플리케이션명·세부업무명을 떼어 낸 글자열.

    단계마다 이름 붙이는 버릇이 달라 한쪽만 접두를 붙이는 일이 잦다.
      설계단계 : "- 사용자 등록"
      종료단계 : "사용자관리 - 사용자 등록"
    이대로 견주면 유사도가 0.83 밖에 안 나오지만, 접두를 떼면 같은 이름이 된다.
    """
    p = norm(row.get("proc"), opts)
    for f in ("biz", "app"):
        head = norm(row.get(f), opts)
        if head and p.startswith(head) and len(p) > len(head):
            p = p[len(head):]
    return p


def diff_of(a, b, fields):
    out = []
    for k, label in DIFF_FIELDS:
        if k not in fields:
            continue
        va, vb = a.get(k), b.get(k)
        if k in ("wt", "ftr", "det", "chg"):
            na, nb = _num(va), _num(vb)
            same = (na == nb)
        else:
            same = (U._s(va).upper() == U._s(vb).upper())
        if not same:
            out.append({"field": k, "label": label,
                        "a": "" if va is None else U._s(va),
                        "b": "" if vb is None else U._s(vb)})
    return out


def compare(group_a, group_b, opts=None, progress=None, cancel=None, on_stage=None):
    o = dict(DEFAULT_OPTS)
    o.update(opts or {})
    diff_fields = set(o.get("diff") or [])

    A, B = group_a["rows"], group_b["rows"]
    for i, r in enumerate(A):
        r["_i"] = i
        r["_k"] = row_key(r, o)
    for i, r in enumerate(B):
        r["_i"] = i
        r["_k"] = row_key(r, o)

    # 같은 그룹 안의 중복. 짝짓기 기준과 따로 본다.
    dfields, dlabel = dup_fields(o)

    def dup_map(rows, fields):
        seen = {}
        for r in rows:
            seen.setdefault(row_key(r, o, fields), []).append(r)
        return {k: v for k, v in seen.items() if len(v) > 1 and k.strip("|")}

    dup_a, dup_b = dup_map(A, dfields), dup_map(B, dfields)
    # 짝짓기 기준만으로 셌을 때와 견줘, 이름이 겹치는 파일인지 알려 준다
    key_fields = KEY_FIELDS.get(o.get("key", "app_proc"), KEY_FIELDS["app_proc"])[0]
    loose = 0
    if list(key_fields) != list(dfields):
        loose = sum(len(v) for v in dup_map(A, key_fields).values()) + \
                sum(len(v) for v in dup_map(B, key_fields).values())
    tight = sum(len(v) for v in dup_a.values()) + sum(len(v) for v in dup_b.values())
    dup_note = ""
    if loose > tight + 20 and "desc" not in key_fields:
        dup_note = ("짝짓기 기준(%s)만으로는 %d건이 같은 이름으로 잡힙니다. "
                    "단위프로세스명이 화면 묶음 이름이고 실제 기능은 설명에 적혀 있는 자료로 보입니다. "
                    "짝짓기 기준에도 설명을 넣으면 짝이 정확해집니다."
                    % (KEY_FIELDS[o.get("key", "app_proc")][1], loose))

    idx_b = {}
    for r in B:
        idx_b.setdefault(r["_k"], []).append(r)

    _tick(progress, "이름이 같은 것 찾는 중", 0, len(A), "")
    pairs = []
    used_b = set()
    left_a = []
    for n_done, r in enumerate(A):
        if cancel and cancel() and n_done % 200 == 0:
            raise Cancelled()
        if n_done % 500 == 0:
            _tick(progress, "이름이 같은 것 찾는 중", n_done, len(A), "")
        cand = [x for x in idx_b.get(r["_k"], []) if x["_i"] not in used_b]
        if not r["_k"].strip("|") or not cand:
            left_a.append(r)
            continue
        m = cand[0]
        used_b.add(m["_i"])
        d = diff_of(r, m, diff_fields)
        pairs.append({"status": "변경" if d else "일치", "score": 1.0,
                      "a": r, "b": m, "diffs": d})

    left_b = [x for x in B if x["_i"] not in used_b]

    if on_stage:
        rest = [{"status": "미확인", "score": 0, "a": r, "b": None, "diffs": []}
                for r in left_a]
        rest += [{"status": "미확인", "score": 0, "a": None, "b": r, "diffs": []}
                 for r in left_b]
        on_stage(build(o, group_a, group_b, pairs + rest, dup_a, dup_b, False,
                       dlabel, dup_note))

    # 이름이 비슷한 것 찾기
    # 남은 것끼리 전부 견주면 건수의 곱만큼 걸리므로,
    # difflib 의 값싼 어림 비교로 먼저 걸러 낸 뒤 정확히 잰다.
    if o.get("fuzzy") and left_a and left_b:
        th = float(o.get("threshold", 0.85))
        total = len(left_a)
        _tick(progress, "이름이 비슷한 것 찾는 중", 0, total,
              "%d × %d 건" % (total, len(left_b)))
        texts_b = [(x, row_text(x, o), proc_text(x, o)) for x in left_b]
        texts_b = [(x, t, p) for x, t, p in texts_b if t]
        sm = difflib.SequenceMatcher(autojunk=False)      # 전체 이름
        smp = difflib.SequenceMatcher(autojunk=False)     # 접두 뗀 단위프로세스명
        still_a, taken = [], set()
        for n_done, r in enumerate(left_a):
            if cancel and cancel():
                raise Cancelled()
            if n_done % 50 == 0:
                _tick(progress, "이름이 비슷한 것 찾는 중", n_done, total,
                      U._s(r.get("proc"))[:30])
            ta, pa = row_text(r, o), proc_text(r, o)
            best, best_s, best_by = None, 0.0, ""
            if ta:
                sm.set_seq2(ta)
                if pa:
                    smp.set_seq2(pa)
                for x, tb, pb in texts_b:
                    if x["_i"] in taken:
                        continue
                    s2, by = 0.0, "이름 전체"
                    if pa and pb:
                        smp.set_seq1(pb)
                        if smp.real_quick_ratio() >= th and smp.quick_ratio() >= th:
                            s2 = smp.ratio()
                            by = "단위프로세스명(접두 제외)"
                    if s2 < th:
                        sm.set_seq1(tb)
                        if sm.real_quick_ratio() >= th and sm.quick_ratio() >= th:
                            v = sm.ratio()
                            if v > s2:
                                s2, by = v, "이름 전체"
                    if s2 > best_s:
                        best, best_s, best_by = x, s2, by
                        if best_s >= 0.999:
                            break
            if best is not None and best_s >= th:
                taken.add(best["_i"])
                pairs.append({"status": "유사", "score": round(best_s, 4),
                              "basis": best_by, "a": r, "b": best,
                              "diffs": diff_of(r, best, diff_fields)})
            else:
                still_a.append(r)
        left_a = still_a
        left_b = [x for x in left_b if x["_i"] not in taken]
        _tick(progress, "이름이 비슷한 것 찾는 중", total, total, "")

    for r in left_a:
        pairs.append({"status": "A만 있음", "score": 0, "a": r, "b": None, "diffs": []})
    for r in left_b:
        pairs.append({"status": "B만 있음", "score": 0, "a": None, "b": r, "diffs": []})

    _tick(progress, "정리하는 중", 1, 1, "")
    order = {"A만 있음": 0, "B만 있음": 1, "변경": 2, "유사": 3, "일치": 4, "미확인": 5}
    pairs.sort(key=lambda p: (order.get(p["status"], 9),
                              U._s((p["a"] or p["b"]).get("app")),
                              U._s((p["a"] or p["b"]).get("proc"))))

    return build(o, group_a, group_b, pairs, dup_a, dup_b, True,
                 dlabel, dup_note)


def build(o, group_a, group_b, pairs, dup_a, dup_b, done,
          dup_label="", dup_note=""):
    """화면·JSON 에 넘길 결과 꾸러미. 진행 중 스냅숏도 같은 모양이다."""
    return {"opts": o, "done": done,
            "a": summarize(group_a), "b": summarize(group_b),
            "pairs": [strip_pair(p) for p in pairs],
            "stats": stats_of(pairs, dup_a, dup_b),
            "keyLabel": KEY_FIELDS.get(o.get("key", "app_proc"),
                                       KEY_FIELDS["app_proc"])[1],
            "dupLabel": dup_label, "dupNote": dup_note,
            "dupA": [{"key": k, "rows": [brief(x) for x in v]} for k, v in dup_a.items()],
            "dupB": [{"key": k, "rows": [brief(x) for x in v]} for k, v in dup_b.items()]}


def brief(r):
    return {"src": r.get("src"), "no": r.get("no"), "app": r.get("app"),
            "biz": r.get("biz"), "proc": r.get("proc"), "type": r.get("type"),
            "cx": r.get("cx"), "wt": r.get("wt"), "ftr": r.get("ftr"),
            "det": r.get("det"), "dev": r.get("dev"), "desc": r.get("desc")}


def strip_pair(p):
    return {"status": p["status"], "score": p["score"], "diffs": p["diffs"],
            "basis": p.get("basis", ""),
            "a": brief(p["a"]) if p["a"] else None,
            "b": brief(p["b"]) if p["b"] else None}


def summarize(g):
    by = {}
    fp = 0.0
    for r in g["rows"]:
        t = r.get("type") or "(없음)"
        b = by.setdefault(t, {"n": 0, "fp": 0.0})
        b["n"] += 1
        if r.get("wt"):
            b["fp"] += r["wt"]
            fp += r["wt"]
    return {"label": g["label"], "files": g["files"], "errors": g["errors"],
            "count": len(g["rows"]), "fp": round(fp, 2), "byType": by}


def stats_of(pairs, dup_a, dup_b):
    s = {"일치": 0, "변경": 0, "유사": 0, "A만 있음": 0, "B만 있음": 0}
    fp_a = fp_b = 0.0
    for p in pairs:
        s[p["status"]] = s.get(p["status"], 0) + 1
        if p["a"] and p["a"].get("wt"):
            fp_a += p["a"]["wt"]
        if p["b"] and p["b"].get("wt"):
            fp_b += p["b"]["wt"]
    s["중복A"] = sum(len(v) for v in dup_a.values())
    s["중복B"] = sum(len(v) for v in dup_b.values())
    s["fpA"] = round(fp_a, 2)
    s["fpB"] = round(fp_b, 2)
    s["fpDiff"] = round(fp_b - fp_a, 2)
    return s


# ---------------------------------------------------------------- JSON

def save_json(result, path):
    """결과(또는 진행 중 스냅숏)를 JSON 으로 남긴다. 오래 걸리는 대조를 잃지 않기 위함."""
    import json
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    os.replace(tmp, path)          # 쓰다 만 파일이 남지 않게
    return path


def load_json(path):
    import json
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------- 내보내기

FONT = "맑은 고딕"

STATUS_COLOR = {                      # 글자색, 채움색
    "일치":     ("375623", "E2EFDA"),
    "변경":     ("BF8F00", "FFF2CC"),
    "유사":     ("6A1B9A", "EDE7F6"),
    "A만 있음": ("C62828", "FDECEA"),
    "B만 있음": ("1565C0", "E3F2FD"),
    "미확인":   ("808080", "F2F2F2"),
}
A_FILL = "DDEBF7"      # A 그룹 열 머리
B_FILL = "FCE4D6"      # B 그룹 열 머리


def _st(ws, cell, *, bold=False, size=10, color="000000", fill=None,
        align="left", wrap=False, border=True, fmt=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        thin = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if fmt:
        cell.number_format = fmt
    return cell


def _edge(ws, row, col, side="left"):
    """A 그룹과 B 그룹 사이에 굵은 세로줄을 넣는다."""
    from openpyxl.styles import Border, Side
    c = ws.cell(row, col)
    thin = Side(style="thin", color="BFBFBF")
    med = Side(style="medium", color="808080")
    b = c.border
    c.border = Border(left=med if side == "left" else (b.left or thin),
                      right=med if side == "right" else (b.right or thin),
                      top=b.top or thin, bottom=b.bottom or thin)


def export_xlsx(result, path):
    """대조 결과를 눈으로 보기 좋게 엑셀로 낸다.

    - A 그룹과 B 그룹을 좌우로 나누고 머리 색과 굵은 세로줄로 가른다
    - 상태별로 줄 색을 달리하고, 달라진 값은 빨간 글씨로 적는다
    - 요약·유형별 집계는 값을 박아 넣지 않고 수식으로 걸어, 엑셀에서 필터를 바꾸면 따라 움직인다
    - 상태별 건수와 유형별 기능점수는 막대그래프로 그린다
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    a_lab, b_lab = result["a"]["label"], result["b"]["label"]
    pairs = result["pairs"]

    # 중복 점검에서 걸린 행을 찾아 두었다가 차이 칸에 적는다
    dup_idx = {}
    for side, lab, groups in (("a", a_lab, result.get("dupA") or []),
                              ("b", b_lab, result.get("dupB") or [])):
        for g in groups:
            grp = g.get("rows") or []
            for r0 in grp:
                others = [str(x["no"]) for x in grp if x["no"] != r0["no"]]
                dup_idx[(side, r0.get("src"), r0.get("no"))] = (lab, len(grp), others)

    def notes_of(p):
        out = []
        for x in p["diffs"]:
            out.append("%s %s→%s" % (x["label"], x["a"] or "-", x["b"] or "-"))
        if p["status"] == "유사" and p.get("basis"):
            out.append("유사 판정 근거 : %s (%.3f)" % (p["basis"], p["score"]))
        for side in ("a", "b"):
            row = p.get(side)
            if not row:
                continue
            hit = dup_idx.get((side, row.get("src"), row.get("no")))
            if hit:
                lab, n, others = hit
                tail = ", ".join(others[:8]) + ("…" if len(others) > 8 else "")
                out.append("%s 중복 %d건 (같은 내용 %s행)" % (lab, n, tail))
        return out
    wb = Workbook()

    def pair_sheet(ws, title, plist, subtitle, tab):
        """대조표 한 장을 그린다. 줄 세우는 차례만 다르고 모양은 같다."""
        ws.title = title
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:S1")
        _st(ws, ws["A1"], bold=True, size=14, align="center", border=False)
        ws["A1"] = "FP 산정 결과물 대조   %s  ↔  %s" % (a_lab, b_lab)
        ws.merge_cells("A2:S2")
        _st(ws, ws["A2"], size=9, color="595959", align="center", border=False)
        ws["A2"] = subtitle

        ws.merge_cells("A3:B3")
        _st(ws, ws["A3"], bold=True, align="center", fill="D9D9D9")
        ws["A3"] = "판정"
        ws.merge_cells("C3:J3")
        _st(ws, ws["C3"], bold=True, align="center", fill=A_FILL)
        ws["C3"] = a_lab
        ws.merge_cells("K3:R3")
        _st(ws, ws["K3"], bold=True, align="center", fill=B_FILL)
        ws["K3"] = b_lab
        _st(ws, ws["S3"], bold=True, align="center", fill="D9D9D9")
        ws["S3"] = "차이"

        cols = ["상태", "유사도"] + \
               ["파일", "행", "애플리케이션", "세부업무", "단위프로세스명",
                "FP유형", "복잡도", "가중치"] * 2 + \
               ["달라진 항목 · 비고"]
        for i2, name in enumerate(cols, start=1):
            fill = A_FILL if 3 <= i2 <= 10 else (B_FILL if 11 <= i2 <= 18 else "D9D9D9")
            _st(ws, ws.cell(4, i2), bold=True, align="center", fill=fill, wrap=True)
            ws.cell(4, i2).value = name

        r = 5
        for p in plist:
            fg, bg = STATUS_COLOR.get(p["status"], ("000000", "FFFFFF"))
            _st(ws, ws.cell(r, 1), bold=True, align="center", color=fg, fill=bg)
            ws.cell(r, 1).value = p["status"].replace("A만", a_lab + "만").replace("B만", b_lab + "만")
            _st(ws, ws.cell(r, 2), align="center", fmt="0.000")
            if p["status"] == "유사":
                ws.cell(r, 2).value = p["score"]

            for base, side in ((3, "a"), (11, "b")):
                row = p[side]
                tint = None if p["status"] == "일치" else bg
                for k in range(8):
                    _st(ws, ws.cell(r, base + k), fill=tint,
                        align="center" if k in (1, 5, 6, 7) else "left")
                if row:
                    ws.cell(r, base).value = row["src"]
                    ws.cell(r, base + 1).value = row["no"]
                    ws.cell(r, base + 2).value = row["app"]
                    ws.cell(r, base + 3).value = row["biz"]
                    ws.cell(r, base + 4).value = row["proc"]
                    ws.cell(r, base + 5).value = row["type"]
                    ws.cell(r, base + 6).value = row["cx"]
                    ws.cell(r, base + 7).value = row["wt"]
                _edge(ws, r, base, "left")

            notes = notes_of(p)
            has_diff = bool(p["diffs"])
            _st(ws, ws.cell(r, 19),
                color="C00000" if has_diff else ("BF8F00" if notes else "000000"),
                wrap=True, fill=bg if notes else None)
            ws.cell(r, 19).value = "\n".join(notes)
            r += 1

        last_r = r - 1
        ws.freeze_panes = "C5"
        if last_r >= 5:
            ws.auto_filter.ref = "A4:S%d" % last_r
        for col, w in zip(range(1, 20),
                          [18, 8, 20, 6, 15, 13, 32, 8, 8, 8,
                           20, 6, 15, 13, 32, 8, 8, 8, 34]):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[4].height = 28
        ws.sheet_properties.tabColor = tab
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "3:4"          # 인쇄할 때 머리글 되풀이
        return last_r

    head = ("짝짓기 기준 : %s   |   유사도 임계값 %s   |   %s 기능 %d건 %.1fFP"
            "   |   %s 기능 %d건 %.1fFP   |   작성 %s"
            % (result["keyLabel"], result["opts"].get("threshold"),
               a_lab, result["a"]["count"], result["a"]["fp"],
               b_lab, result["b"]["count"], result["b"]["fp"],
               time.strftime("%Y-%m-%d %H:%M")))

    # ────────────────────────────────── 대조결과 (상태순)
    ws = wb.active
    last = pair_sheet(ws, "대조결과", pairs,
                      head + "   |   상태순", "37474F")

    # ──────────────── 대조결과(A그룹기준) : A 그룹 파일·행 차례 그대로
    order = {f["name"]: i for i, f in enumerate(result["a"]["files"])}
    border = {f["name"]: i for i, f in enumerate(result["b"]["files"])}
    with_a = sorted([p for p in pairs if p["a"]],
                    key=lambda p: (order.get(p["a"]["src"], 9999), p["a"]["no"]))
    only_b = sorted([p for p in pairs if not p["a"]],
                    key=lambda p: (border.get(p["b"]["src"], 9999), p["b"]["no"]))
    pair_sheet(wb.create_sheet("대조결과(A그룹기준)"), "대조결과(A그룹기준)",
               with_a + only_b,
               head + "   |   %s 파일·행 차례 그대로 (%s에만 있는 %d건은 맨 아래)"
               % (a_lab, b_lab, len(only_b)), "1565C0")

    n0, n1 = 5, max(last, 5)

    def cnt(status):
        return '=COUNTIF(대조결과!$A$%d:$A$%d,"*%s*")' % (n0, n1, status)

    # ────────────────────────────────── 요약
    s = wb.create_sheet("요약")
    s.sheet_view.showGridLines = False
    s.sheet_properties.tabColor = "1565C0"
    s.merge_cells("A1:E1")
    _st(s, s["A1"], bold=True, size=14, align="center", border=False)
    s["A1"] = "대조 요약"

    s["A3"] = "구분"
    s["B3"] = a_lab
    s["C3"] = b_lab
    s["D3"] = "차이 (%s − %s)" % (b_lab, a_lab)
    for c in "ABCD":
        _st(s, s[c + "3"], bold=True, align="center", fill="D9D9D9")
    rows = [("파일 수", len(result["a"]["files"]), len(result["b"]["files"]), "#,##0"),
            ("기능 수", '=COUNT(대조결과!$D$%d:$D$%d)' % (n0, n1),
             '=COUNT(대조결과!$L$%d:$L$%d)' % (n0, n1), "#,##0"),
            ("기능점수", '=SUM(대조결과!$J$%d:$J$%d)' % (n0, n1),
             '=SUM(대조결과!$R$%d:$R$%d)' % (n0, n1), "#,##0.0")]
    rr = 4
    for label, av, bv, fmt in rows:
        _st(s, s.cell(rr, 1), bold=True, fill="F2F2F2")
        s.cell(rr, 1).value = label
        for col, v in ((2, av), (3, bv)):
            _st(s, s.cell(rr, col), align="right", fmt=fmt)
            s.cell(rr, col).value = v
        _st(s, s.cell(rr, 4), align="right", bold=True,
            fmt="+%s;-%s;0" % (fmt, fmt))
        s.cell(rr, 4).value = "=C%d-B%d" % (rr, rr)
        rr += 1

    rr += 1
    s.cell(rr, 1).value = "상태"
    s.cell(rr, 2).value = "건수"
    for c in (1, 2):
        _st(s, s.cell(rr, c), bold=True, align="center", fill="D9D9D9")
    head = rr
    rr += 1
    for st in ("일치", "변경", "유사", "A만 있음", "B만 있음", "미확인"):
        fg, bg = STATUS_COLOR[st]
        _st(s, s.cell(rr, 1), bold=True, color=fg, fill=bg)
        s.cell(rr, 1).value = st.replace("A만", a_lab + "만").replace("B만", b_lab + "만")
        _st(s, s.cell(rr, 2), align="right", fmt="#,##0")
        s.cell(rr, 2).value = cnt(st.replace("A만 있음", "만 있음").replace("B만 있음", "만 있음")
                                  if st.endswith("만 있음") else st)
        rr += 1
    # A만/B만 은 그룹 이름이 앞에 붙으므로 정확한 조건으로 다시 건다
    s.cell(head + 4, 2).value = '=COUNTIF(대조결과!$A$%d:$A$%d,"%s만 있음")' % (n0, n1, a_lab)
    s.cell(head + 5, 2).value = '=COUNTIF(대조결과!$A$%d:$A$%d,"%s만 있음")' % (n0, n1, b_lab)
    last_st = rr - 1

    ch = BarChart()
    ch.type = "bar"
    ch.title = "상태별 건수"
    ch.height, ch.width = 7, 12
    ch.legend = None
    data = Reference(s, min_col=2, min_row=head, max_row=last_st)
    cats = Reference(s, min_col=1, min_row=head + 1, max_row=last_st)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    s.add_chart(ch, "F3")

    for col, w in (("A", 22), ("B", 14), ("C", 14), ("D", 20)):
        s.column_dimensions[col].width = w

    note = last_st + 2
    _st(s, s.cell(note, 1), size=9, color="595959", border=False)
    s.cell(note, 1).value = ("건수·기능점수는 대조결과 시트를 가리키는 수식입니다. "
                             "시트에서 행을 지우면 이 값도 따라 바뀝니다.")

    # ────────────────────────────────── 유형별
    t = wb.create_sheet("유형별")
    t.sheet_view.showGridLines = False
    t.sheet_properties.tabColor = "2E7D32"
    t.merge_cells("A1:F1")
    _st(t, t["A1"], bold=True, size=14, align="center", border=False)
    t["A1"] = "FP유형별 대비"
    heads = ["FP유형", "%s 건수" % a_lab, "%s FP" % a_lab,
             "%s 건수" % b_lab, "%s FP" % b_lab, "FP 차이"]
    for i, h in enumerate(heads, start=1):
        fill = A_FILL if i in (2, 3) else (B_FILL if i in (4, 5) else "D9D9D9")
        _st(t, t.cell(3, i), bold=True, align="center", fill=fill)
        t.cell(3, i).value = h
    types = sorted(set(list(result["a"]["byType"]) + list(result["b"]["byType"])))
    rr = 4
    for ty in types:
        _st(t, t.cell(rr, 1), bold=True, align="center")
        t.cell(rr, 1).value = ty
        t.cell(rr, 2).value = '=COUNTIF(대조결과!$H$%d:$H$%d,$A%d)' % (n0, n1, rr)
        t.cell(rr, 3).value = '=SUMIF(대조결과!$H$%d:$H$%d,$A%d,대조결과!$J$%d:$J$%d)' % (n0, n1, rr, n0, n1)
        t.cell(rr, 4).value = '=COUNTIF(대조결과!$P$%d:$P$%d,$A%d)' % (n0, n1, rr)
        t.cell(rr, 5).value = '=SUMIF(대조결과!$P$%d:$P$%d,$A%d,대조결과!$R$%d:$R$%d)' % (n0, n1, rr, n0, n1)
        t.cell(rr, 6).value = "=E%d-C%d" % (rr, rr)
        for c in range(2, 7):
            _st(t, t.cell(rr, c), align="right", fmt="#,##0.0",
                fill=A_FILL if c in (2, 3) else (B_FILL if c in (4, 5) else None))
        _st(t, t.cell(rr, 6), align="right", bold=True, fmt="+#,##0.0;-#,##0.0;0")
        rr += 1
    tlast = rr - 1
    for col, w in zip("ABCDEF", (12, 14, 14, 14, 14, 14)):
        t.column_dimensions[col].width = w

    if tlast >= 4:
        ch2 = BarChart()
        ch2.type = "col"
        ch2.grouping = "clustered"
        ch2.title = "FP유형별 기능점수"
        ch2.height, ch2.width = 8, 14
        d2 = Reference(t, min_col=3, max_col=5, min_row=3, max_row=tlast)
        c2 = Reference(t, min_col=1, min_row=4, max_row=tlast)
        ch2.add_data(d2, titles_from_data=True)
        ch2.set_categories(c2)
        ch2.series = [ch2.series[0], ch2.series[2]]      # 건수 열은 빼고 FP 만
        t.add_chart(ch2, "H3")

    # ────────────────────────────────── 읽은 파일
    f = wb.create_sheet("읽은 파일")
    f.sheet_view.showGridLines = False
    for i, h in enumerate(["그룹", "파일", "양식", "시트", "산정방법", "기능 수"], start=1):
        _st(f, f.cell(1, i), bold=True, align="center", fill="D9D9D9")
        f.cell(1, i).value = h
    rr = 2
    for g, lab, fill in ((result["a"], a_lab, A_FILL), (result["b"], b_lab, B_FILL)):
        for x in g["files"]:
            vals = [lab, x["name"], x["formName"], x["sheet"], x["method"], x["count"]]
            for i, v in enumerate(vals, start=1):
                _st(f, f.cell(rr, i), fill=fill if i == 1 else None,
                    align="right" if i == 6 else "left")
                f.cell(rr, i).value = v
            rr += 1
        for e in g.get("errors", []):
            _st(f, f.cell(rr, 1), fill=fill)
            f.cell(rr, 1).value = lab
            _st(f, f.cell(rr, 2), color="C00000")
            f.cell(rr, 2).value = "%s — %s" % (e["path"], e["msg"])
            rr += 1
    for col, w in zip("ABCDEF", (14, 40, 34, 16, 12, 10)):
        f.column_dimensions[col].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------- 단독 실행

def _cli():
    args = sys.argv[1:]
    if "--" not in args:
        print(__doc__)
        return 1
    i = args.index("--")
    fa, fb = args[:i], args[i + 1:]
    if not fa or not fb:
        print("양쪽 파일을 모두 지정하십시오.")
        return 1
    ga = load_group(fa, "A")
    gb = load_group(fb, "B")
    r = compare(ga, gb)
    st = r["stats"]
    print("A %d개 파일 / 기능 %d건 / %.1f FP" % (len(ga["files"]), r["a"]["count"], r["a"]["fp"]))
    print("B %d개 파일 / 기능 %d건 / %.1f FP" % (len(gb["files"]), r["b"]["count"], r["b"]["fp"]))
    print("일치 %d · 변경 %d · 유사 %d · A만 %d · B만 %d · FP차이 %+.1f"
          % (st["일치"], st["변경"], st["유사"], st["A만 있음"], st["B만 있음"], st["fpDiff"]))
    for p in r["pairs"]:
        if p["status"] in ("A만 있음", "B만 있음"):
            r0 = p["a"] or p["b"]
            print("  [%s] %s / %s (%s %d행)"
                  % (p["status"], r0["app"], r0["proc"], r0["src"], r0["no"]))
    return 0


if __name__ == "__main__":
    sys.exit(_cli())