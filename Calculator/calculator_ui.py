# -*- coding: utf-8 -*-
"""
calculator_ui.py  v3.0
SW개발비 기능점수(FP) 산정양식 편집기 - 화면·실행 담당

역할 분리
  calculator_ui.py       화면(HTML), 웹서버, 엑셀 읽기·쓰기 엔진, 개발비 계산 모델
  calculator_detail.py   정통법·상세법 양식 사양  (2024-정통법 / 2016-상세법)
  calculator_simple.py   간이법 양식 사양          (2024-간이법)

양식을 새로 추가할 때는 사양 파일에 항목 하나를 더 쓰면 되고 화면은 건드리지 않는다.

실행
  python calculator_ui.py [파일경로] [--form 2024-정통법] [--no-browser]
  python calculator_detail.py   정통법·상세법만 목록에 띄우고 시작
  python calculator_simple.py   간이법만 목록에 띄우고 시작
"""

import os
import re
import sys
import time
import random
import socket
import shutil
import tempfile
import threading
import subprocess
import configparser
import webbrowser

from flask import Flask, request, jsonify, Response

import openpyxl

APP_NAME = "SW개발비 FP 산정양식 편집기"
VERSION = "3.0"

FP_TYPES = ["ILF", "EIF", "EI", "EO", "EQ"]
SIMPLE_WEIGHT = {"ILF": 7.5, "EIF": 5.4, "EI": 4.0, "EO": 5.2, "EQ": 3.9}
DEV_TYPES = ["신규 개발", "수정 후 재사용", "수정 없이 재사용"]

# ---------------------------------------------------------------- 양식 정의
#
# fp.cols   : 화면 열 정의. k=키, x=엑셀 열번호, t=입력형태
#             text 자유입력 / num 숫자 / sel 선택 / ro 계산결과(읽기전용)
# fp.anchor : 머리글 행을 찾는 표식. 앞쪽 작성가이드 본문에 같은 낱말이 나와도
#             애플리케이션명 열과 FP유형 열이 같은 행에 있어야 머리글로 본다.

# 양식 사양은 별도 파일에서 불러온다.
import calculator_detail as SPEC_DETAIL
import calculator_simple as SPEC_SIMPLE

try:
    import calculator_planned as SPEC_PLANNED      # 추후 개발예정 목록(없어도 동작)
except Exception:
    SPEC_PLANNED = None

FORMS = {}
for _mod in (SPEC_DETAIL, SPEC_SIMPLE, SPEC_PLANNED):
    if _mod is None:
        continue
    for _k, _v in _mod.FORMS.items():
        FORMS[_k] = _v


def is_ready(key):
    """실제로 다룰 수 있는 양식인가."""
    return FORMS.get(key, {}).get("status") != "planned"


GROUP_ORDER = ["기획단계", "구현단계", "운영단계", "데이터베이스", "기타"]


def form_list(keys):
    """화면 목록용. 단계 → 양식번호 순으로 정렬한다."""
    def sort_key(k):
        f = FORMS[k]
        g = f.get("group", "기타")
        gi = GROUP_ORDER.index(g) if g in GROUP_ORDER else len(GROUP_ORDER)
        return (gi, f.get("code", ""), f.get("name", ""))
    out = []
    for k in sorted(keys, key=sort_key):
        f = FORMS[k]
        out.append({"key": k, "name": f.get("name", k),
                    "code": f.get("code", ""), "group": f.get("group", "기타"),
                    "file": f.get("file", ""), "ready": is_ready(k)})
    return out


# 화면에 보여줄 양식 목록(제한 가능). main(only=...) 으로 좁힌다.
VISIBLE = list(FORMS)


# 2024 양식 개발비 시트 입력 셀
C24 = {"unitPrice": "D7", "profitRate": "J9",
       "adj": {"link": "D22", "perf": "D23", "env": "D24", "sec": "D25"},
       "expenseRows": [15, 16]}

ADJ_TABLE = {
    "link": [
        ["1. 타 기관 연계 없음", 0.88],
        ["2. 1~2개의 타 기관 연계", 0.94],
        ["3. 3~5개의 타 기관 연계", 1.0],
        ["4. 6~10개의 타 기관 연계", 1.06],
        ["5. 10개 초과의 타 기관 연계", 1.12],
    ],
    "perf": [
        ["1. 응답성능에 대한 특별한 요구사항이 없다.", 0.91],
        ["2. 응답성능에 대한 요구사항이 있으나 특별한 조치가 필요하지는 않다.", 0.95],
        ["3. 응답시간이나 처리율이 피크타임(peak time)에 중요하며, 처리 시한이 명시되어 있다.", 1.0],
        ["4. 응답시간이나 처리율이 모든 업무시간에 중요하며, 처리 시한이 명시되어 있다.", 1.05],
        ["5. 응답성능 요구수준이 엄격하여, 설계, 개발 또는 구현 단계에서 성능 분석도구 사용이 필요하다.", 1.09],
    ],
    "env": [
        ["1. 운영환경 호환성에 대한 요구사항이 없다.", 0.94],
        ["2. 운영환경 호환성에 대한 요구사항이 있으며, 동일 하드웨어 및 소프트웨어 환경에서 운영되도록 설계된다.", 1.0],
        ["3. 운영환경 호환성에 대한 요구사항이 있으며, 유사 하드웨어 및 소프트웨어 환경에서 운영되도록 설계된다.", 1.06],
        ["4. 운영환경 호환성에 대한 요구사항이 있으며, 이질적인 하드웨어 및 소프트웨어 환경에서 운영되도록 설계된다.", 1.13],
        ["5. 항목 4에 더하여 일반적 산출물 이외에 장소에서 원활한 운영을 보장하기 위한 운영 절차의 문서화와 사전 모의훈련이 요구된다.", 1.19],
    ],
    "sec": [
        ["1. 암호화, 웹취약점 점검, 시큐어코딩, 개인정보보호 등 1가지 보안 요구사항이 포함되어 있다.", 0.97],
        ["2. 2가지 요구사항이 포함되어 있다.", 1.0],
        ["3. 3가지 요구사항이 포함되어 있다.", 1.03],
        ["4. 4가지 항목이 모두 포함되어 있다.", 1.06],
        ["5. 5가지 이상의 보안 요구사항이 포함되어 있다.", 1.08],
    ],
}
ADJ_LABEL = {"link": "연계복잡성", "perf": "성능 요구수준",
             "env": "운영환경 호환성", "sec": "보안성 요구수준"}


# ---------------------------------------------------------------- 공통 유틸

def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


INI_PATH = os.path.join(app_dir(), "calculator_ui.ini")


def load_ini():
    cp = configparser.ConfigParser()
    cp.read(INI_PATH, encoding="utf-8")
    if "main" not in cp:
        cp["main"] = {}
    return cp


def save_ini(cp):
    try:
        with open(INI_PATH, "w", encoding="utf-8") as f:
            cp.write(f)
    except Exception:
        pass


def setup_output():
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w", encoding="utf-8")


def _no_window():
    if os.name == "nt":
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        return {"startupinfo": si, "creationflags": 0x08000000}
    return {}


def pick_port():
    for _ in range(200):
        p = random.randint(49152, 65535)
        s = socket.socket()
        try:
            s.bind(("127.0.0.1", p))
            s.close()
            return p
        except OSError:
            try:
                s.close()
            except Exception:
                pass
    return 8731


def find_chrome():
    cands = []
    if os.name == "nt":
        for base in [os.environ.get("PROGRAMFILES", r"C:\Program Files"),
                     os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"),
                     os.environ.get("LOCALAPPDATA", "")]:
            if not base:
                continue
            cands += [os.path.join(base, r"Google\Chrome\Application\chrome.exe"),
                      os.path.join(base, r"Microsoft\Edge\Application\msedge.exe")]
    else:
        cands += ["/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
                  shutil.which("google-chrome") or "",
                  shutil.which("chromium") or ""]
    for c in cands:
        if c and os.path.exists(c):
            return c
    return None


def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _n(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _flat(v):
    return re.sub(r"[\s\u3000①-⑳*·:()\[\]]", "", _s(v))


_ROWREF = re.compile(r"(\$?[A-Za-z]{1,3})(\$?)(\d+)(?!\()")


def shift_formula(f, src, dst):
    """수식의 상대 행참조만 src -> dst 로 바꾼다. $행 은 그대로 둔다."""
    def rep(m):
        col, dollar, num = m.groups()
        if dollar != "$" and int(num) == src:
            return "%s%s%d" % (col, dollar, dst)
        return m.group(0)
    return _ROWREF.sub(rep, f)


def coef_of(label):
    """'업무처리용 (1.0)' -> 1.0"""
    m = re.findall(r"\(([\d.]+)\)", _s(label))
    return float(m[-1]) if m else None


# ---------------------------------------------------------------- FP 계산
# 두 양식의 VBA Module1 (FPR / FPV) 을 그대로 옮긴 것

def fp_rating(ttype, ftr, det):
    r = None
    if ttype in ("EO", "EQ"):
        if det == 0 and ftr == 0:
            r = 0
        if 0 < det < 20 and ftr < 2:
            r = "L"
        if det > 19 and ftr < 2:
            r = "A"
        if ftr in (2, 3) and det <= 5:
            r = "L"
        if ftr in (2, 3) and 5 < det < 20:
            r = "A"
        if ftr in (2, 3) and det > 19:
            r = "H"
        if ftr > 3 and det < 6:
            r = "A"
        if ftr > 3 and det >= 6:
            r = "H"
    elif ttype == "EI":
        if det == 0 and ftr == 0:
            r = 0
        if 0 < det < 16 and ftr < 2:
            r = "L"
        if det > 15 and ftr < 2:
            r = "A"
        if ftr == 2 and det <= 4:
            r = "L"
        if ftr == 2 and 4 < det < 16:
            r = "A"
        if ftr == 2 and det >= 16:
            r = "H"
        if ftr > 2 and det < 5:
            r = "A"
        if ftr > 2 and det >= 5:
            r = "H"
    elif ttype in ("ILF", "EIF"):
        if det == 0 and ftr == 0:
            r = 0
        if 0 < det < 51 and ftr < 2:
            r = "L"
        if det > 50 and ftr < 2:
            r = "A"
        if 2 <= ftr <= 5 and det < 20:
            r = "L"
        if 2 <= ftr <= 5 and 19 < det < 51:
            r = "A"
        if 2 <= ftr <= 5 and det >= 51:
            r = "H"
        if ftr >= 6 and det <= 19:
            r = "A"
        if ftr >= 6 and det > 19:
            r = "H"
    return r


FPV_TABLE = {"EI": {"L": 3, "A": 4, "H": 6},
             "EQ": {"L": 3, "A": 4, "H": 6},
             "EO": {"L": 4, "A": 5, "H": 7},
             "ILF": {"L": 7, "A": 10, "H": 15},
             "EIF": {"L": 5, "A": 7, "H": 10}}


def fp_value(ttype, rating):
    return FPV_TABLE.get(ttype, {}).get(rating)


# ---------------------------------------------------------------- 엑셀 읽기

STATE = {"path": None, "wb": None, "form": None, "fp": None, "cost": None,
         "rows": [], "params": {}, "last_row": 0, "payload": None}


def guess_form(path):
    """시트 구성만 보고 어느 양식인지 추정한다(화면 기본 선택값)."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        names = [n.replace(" ", "") for n in wb.sheetnames]
        wb.close()
    except Exception:
        return None
    if any("FP산정(간이법)" in n for n in names):
        return "2024-간이법"
    if any("FP집계" in n for n in names) or any("개발비산출" in n for n in names):
        return "2016-상세법"
    if any("SW개발비산정" in n for n in names):
        return "2024-정통법"
    return None


def pick_sheet(wb, spec):
    for n in spec.get("sheets", []):
        if n in wb.sheetnames:
            return wb[n]
    for key in spec.get("sheet_like", []):
        for n in wb.sheetnames:
            if key in n.replace(" ", ""):
                return wb[n]
    return None


def find_header_row(ws, anchor):
    (ac, akeys) = anchor["app"]
    (tc, tkeys) = anchor["type"]
    limit = min(ws.max_row, 400)
    for r in range(1, limit + 1):
        a = _flat(ws.cell(r, ac).value)
        t = _flat(ws.cell(r, tc).value)
        if not a or not t:
            continue
        if any(k in a for k in akeys) and any(k in t for k in tkeys):
            return r
    return None


def dv_last_row(ws, col):
    best = 0
    try:
        for dv in ws.data_validations.dataValidation:
            for rng in str(dv.sqref).split():
                m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", rng)
                if not m:
                    continue
                c1 = openpyxl.utils.column_index_from_string(m.group(1))
                c2 = openpyxl.utils.column_index_from_string(m.group(3))
                if c1 <= col <= c2:
                    best = max(best, int(m.group(4)))
    except Exception:
        pass
    return best


def find_method_cell(ws, dv_formula):
    try:
        for dv in ws.data_validations.dataValidation:
            if _s(dv.formula1).replace(" ", "") == dv_formula:
                return str(dv.sqref).split()[0].split(":")[0]
    except Exception:
        pass
    return None


def find_label_cell(ws, keys, upto, value_col=5):
    """머리글 앞쪽에서 라벨을 찾아 입력 셀 주소를 돌려준다.
    같은 낱말이 작성가이드 본문에도 나오므로 아래쪽(입력표)부터 거슬러 찾고,
    입력 자리가 병합 셀이면 건너뛴다."""
    letter = openpyxl.utils.get_column_letter(value_col)
    merged = set()
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= value_col <= rng.max_col and rng.min_col != rng.max_col:
            for rr in range(rng.min_row, rng.max_row + 1):
                merged.add(rr)
    for r in range(upto - 1, 0, -1):
        if r in merged:
            continue
        for c in range(2, 6):
            t = _flat(ws.cell(r, c).value)
            if t and any(_flat(k) in t for k in keys):
                return "%s%d" % (letter, r)
    return None


def read_workbook(path, form_key):
    if form_key not in FORMS:
        raise ValueError("알 수 없는 양식 구분입니다: %s" % form_key)
    if not is_ready(form_key):
        raise ValueError("아직 개발하지 않은 양식입니다: %s"
                         % FORMS[form_key].get("name", form_key))
    form = FORMS[form_key]
    keep_vba = path.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=keep_vba)

    ws = pick_sheet(wb, form["fp"])
    if ws is None:
        raise ValueError("%s 에 필요한 FP 시트(%s)가 없습니다. "
                         "양식 선택이 맞는지 확인하십시오."
                         % (form["name"], " / ".join(form["fp"]["sheets"])))
    hrow = find_header_row(ws, form["fp"]["anchor"])
    if hrow is None:
        raise ValueError("'%s' 시트에서 표 머리글(애플리케이션명·FP유형)을 찾지 못했습니다. "
                         "양식 선택이 맞는지 확인하십시오." % ws.title)
    start = hrow + 1
    cols = form["fp"]["cols"]
    cx = {c["k"]: c["x"] for c in cols}
    last_allowed = dv_last_row(ws, cx["type"]) or ws.max_row

    method = form["fp"].get("method", "상세법")
    method_cell = None
    if form["fp"].get("method_dv"):
        method_cell = find_method_cell(ws, form["fp"]["method_dv"])
        if method_cell:
            method = _s(ws[method_cell].value) or method

    ro_keys = [c["k"] for c in cols if c["t"] == "ro"]
    cached = {}
    try:
        wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws2 = wb2[ws.title]
        lo = min(cx[k] for k in ro_keys)
        hi = max(cx[k] for k in ro_keys)
        top = min(last_allowed, ws2.max_row)
        if top >= start:
            for r, vals in enumerate(ws2.iter_rows(min_row=start, max_row=top,
                                                   min_col=lo, max_col=hi,
                                                   values_only=True), start=start):
                if any(v is not None for v in vals):
                    cached[r] = {k: vals[cx[k] - lo] for k in ro_keys}
        wb2.close()
    except Exception:
        cached = {}

    input_keys = [c["k"] for c in cols if c["t"] != "ro"]
    rows = []
    last_row = start - 1
    for r in range(start, min(ws.max_row, last_allowed) + 1):
        rec = {k: _s(ws.cell(r, cx[k]).value) for k in input_keys}
        if not any(rec[k] for k in input_keys):
            continue
        last_row = r
        rec["type"] = rec.get("type", "").upper()
        for k in ro_keys:
            rec["file_" + k] = _s((cached.get(r) or {}).get(k))
        rows.append(rec)

    params = {"method": method, "methodCell": method_cell}

    cost_spec = form["cost"]
    cs = pick_sheet(wb, cost_spec)
    cost = {"kind": "none", "sheet": None}
    if cs is not None:
        ok = all(_flat(cs[k].value).startswith(_flat(v))
                 for k, v in cost_spec.get("check", {}).items())
        if ok:
            cost = {"kind": cost_spec["kind"], "sheet": cs.title}
    if cost["kind"] == "kosa2024":
        cost.update(read_cost_2024(cs))
    elif cost["kind"] == "kosa2016":
        cost.update(read_cost_2016(cs, cost_spec["cells"]))
        cost["rework"] = read_rework(ws, form["fp"]["rework"], hrow)

    payload = {
        "path": path, "name": os.path.basename(path), "form": form_key,
        "formName": form["name"], "sheet": ws.title, "headerRow": hrow,
        "startRow": start, "maxRows": max(0, last_allowed - start + 1),
        "cols": cols, "fpTypes": FP_TYPES, "devTypes": DEV_TYPES,
        "simpleWeight": SIMPLE_WEIGHT, "rows": rows, "params": params,
        "cost": cost, "adjTable": ADJ_TABLE, "adjLabel": ADJ_LABEL,
    }
    STATE.update({"path": path, "wb": wb, "form": form_key,
                  "fp": {"sheet": ws.title, "header": hrow, "start": start,
                         "cx": cx, "cols": cols, "last_allowed": last_allowed},
                  "cost": cost, "rows": rows, "params": params,
                  "last_row": last_row, "payload": payload})
    return payload


def match_option(kind, text):
    table = ADJ_TABLE[kind]
    t = _s(text)
    for label, _v in table:
        if label == t:
            return label
    head = t[:2]
    for label, _v in table:
        if label[:2] == head:
            return label
    return table[0][0]


def read_cost_2024(cs):
    expenses = []
    for r in C24["expenseRows"]:
        expenses.append({"name": _s(cs.cell(r, 2).value),
                         "detail": _s(cs.cell(r, 4).value),
                         "amount": _s(cs.cell(r, 11).value)})
    return {"title": _s(cs["B1"].value),
            "unitPrice": _n(cs[C24["unitPrice"]].value) or 0,
            "profitRate": _n(cs[C24["profitRate"]].value) or 0,
            "adj": {k: match_option(k, cs[c].value) for k, c in C24["adj"].items()},
            "expenses": expenses}


def read_cost_2016(cs, cells):
    stages = []
    for i in range(4):
        stages.append({"name": _s(cs[cells["stageName"][i]].value),
                       "w": _n(cs[cells["stageW"][i]].value) or 0})
    apps, langs, quality = [], [], []
    for lc, sc in zip(cells["appLabel"], cells["appShare"]):
        lab = _s(cs[lc].value)
        if lab:
            apps.append({"label": lab, "coef": coef_of(lab) or 1.0,
                         "share": _n(cs[sc].value) or 0, "cell": sc})
    for lc, sc in zip(cells["langLabel"], cells["langShare"]):
        lab = _s(cs[lc].value)
        if lab:
            langs.append({"label": lab, "coef": coef_of(lab) or 1.0,
                          "share": _n(cs[sc].value) or 0, "cell": sc})
    for lc, vc in zip(cells["qLabel"], cells["qValue"]):
        lab = _s(cs[lc].value)
        if lab:
            quality.append({"label": lab, "val": _n(cs[vc].value) or 0, "cell": vc})
    return {"title": _s(cs["B1"].value),
            "unitPrice": _n(cs[cells["unitPrice"]].value) or 0,
            "profitRate": _n(cs[cells["profitRate"]].value) or 0,
            "stages": stages, "apps": apps, "langs": langs, "quality": quality}


def read_rework(ws, spec, hrow):
    out = {}
    for key, labels, options in spec:
        cell = find_label_cell(ws, labels, hrow)
        out[key] = {"cell": cell, "value": _s(ws[cell].value) if cell else "",
                    "options": options, "label": labels[0]}
    return out


# ---------------------------------------------------------------- 엑셀 쓰기

def template_formula(ws, col, start, last_allowed):
    for r in range(start, min(start + 200, last_allowed) + 1):
        v = ws.cell(r, col).value
        if isinstance(v, str) and v.startswith("="):
            return r, v
    return None, None


def write_workbook(rows, params, out_path):
    wb = STATE["wb"]
    if wb is None:
        raise ValueError("열려 있는 파일이 없습니다.")
    fp = STATE["fp"]
    ws = wb[fp["sheet"]]
    cx, cols, start = fp["cx"], fp["cols"], fp["start"]
    n = len(rows)
    if start + n - 1 > fp["last_allowed"]:
        raise ValueError("양식이 허용하는 행 수(%d행)를 넘었습니다."
                         % (fp["last_allowed"] - start + 1))

    input_keys = [c["k"] for c in cols if c["t"] != "ro"]
    num_keys = {c["k"] for c in cols if c["t"] == "num"}

    # 계산 열 + 보조 열(설계변경률 계산용 P~U 등)의 수식 원본 확보
    targets = [c["x"] for c in cols if c["t"] == "ro"]
    max_x = max(cx.values())
    for c in range(max_x + 1, max_x + 12):
        r0, f0 = template_formula(ws, c, start, fp["last_allowed"])
        if f0:
            targets.append(c)
    tmpl = {}
    for c in targets:
        sr, sf = template_formula(ws, c, start, fp["last_allowed"])
        if sf:
            tmpl[c] = (sr, sf)

    for i, row in enumerate(rows):
        r = start + i
        for k in input_keys:
            v = row.get(k, "")
            if k == "type":
                v = (v or "").upper()
                v = v if v in FP_TYPES else None
            elif k in num_keys:
                v = _n(v)
            else:
                v = v or None
            ws.cell(r, cx[k]).value = v
        for c, (sr, sf) in tmpl.items():
            cur = ws.cell(r, c).value
            if not (isinstance(cur, str) and cur.startswith("=")):
                ws.cell(r, c).value = shift_formula(sf, sr, r)

    for r in range(start + n, max(STATE["last_row"], start + n) + 1):
        for k in input_keys:
            ws.cell(r, cx[k]).value = None

    if params.get("method") and STATE["params"].get("methodCell"):
        ws[STATE["params"]["methodCell"]].value = params["method"]

    cost = STATE["cost"]
    if cost["kind"] == "kosa2024":
        write_cost_2024(wb[cost["sheet"]], params.get("cost", {}))
    elif cost["kind"] == "kosa2016":
        write_cost_2016(wb[cost["sheet"]], ws, params.get("cost", {}))

    wb.save(out_path)
    STATE["path"] = out_path
    STATE["last_row"] = start + n - 1
    return out_path


def write_cost_2024(cs, c):
    if c.get("title"):
        cs["B1"].value = c["title"]
    cs[C24["unitPrice"]].value = _n(c.get("unitPrice")) or 0
    cs[C24["profitRate"]].value = _n(c.get("profitRate")) or 0
    for k, cell in C24["adj"].items():
        cs[cell].value = match_option(k, c.get("adj", {}).get(k))
    for idx, r in enumerate(C24["expenseRows"]):
        lst = c.get("expenses") or []
        e = lst[idx] if idx < len(lst) else {}
        cs.cell(r, 2).value = e.get("name") or None
        cs.cell(r, 4).value = e.get("detail") or None
        cs.cell(r, 11).value = _n(e.get("amount"))


def write_cost_2016(cs, fpws, c):
    cells = FORMS[STATE["form"]]["cost"]["cells"]
    if c.get("title"):
        cs["B1"].value = c["title"]
    cs[cells["unitPrice"]].value = _n(c.get("unitPrice")) or 0
    cs[cells["profitRate"]].value = _n(c.get("profitRate")) or 0
    for grp, key in (("apps", "share"), ("langs", "share"), ("quality", "val")):
        for item in c.get(grp) or []:
            if item.get("cell"):
                cs[item["cell"]].value = _n(item.get(key)) or 0
    for _key, item in (c.get("rework") or {}).items():
        cell = item.get("cell")
        if not cell:
            continue
        v = _s(item.get("value"))
        if v == "(선택)":
            fpws[cell].value = v
        elif _n(v) is not None:
            fpws[cell].value = _n(v)
        else:
            fpws[cell].value = v or None


# ---------------------------------------------------------------- 웹 서버

app = Flask(__name__)
LAST_PING = [time.time()]
CONNECTED = [False]


@app.after_request
def no_cache(resp):
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/")
def index():
    return Response(HTML, mimetype="text/html; charset=utf-8")


@app.route("/api/init")
def api_init():
    cp = load_ini()
    recent = [p for p in cp["main"].get("recent", "").split("|") if p and os.path.exists(p)]
    forms = form_list(VISIBLE)
    return jsonify({"app": APP_NAME, "version": VERSION, "recent": recent,
                    "forms": forms, "data": STATE.get("payload")})


@app.route("/api/guess", methods=["POST"])
def api_guess():
    path = (request.json or {}).get("path", "").strip().strip('"')
    if not path or not os.path.exists(path):
        return jsonify({"ok": False, "msg": "파일을 찾을 수 없습니다."})
    return jsonify({"ok": True, "form": guess_form(path)})


@app.route("/api/open", methods=["POST"])
def api_open():
    body = request.json or {}
    path = body.get("path", "").strip().strip('"')
    if not path:
        return jsonify({"ok": False, "msg": "파일 경로가 비어 있습니다."})
    if not os.path.exists(path):
        return jsonify({"ok": False, "msg": "파일을 찾을 수 없습니다: %s" % path})
    form = body.get("form") or guess_form(path)
    if form and not is_ready(form):
        return jsonify({"ok": False,
                        "msg": "'%s' 은(는) 아직 개발하지 않은 양식입니다."
                               % FORMS[form].get("name", form)})
    if not form:
        ready = [k for k in VISIBLE if is_ready(k)]
        form = ready[0] if ready else ""
    try:
        data = read_workbook(path, form)
    except Exception as e:
        return jsonify({"ok": False, "msg": "열기 실패: %s" % e})
    cp = load_ini()
    recent = [p for p in cp["main"].get("recent", "").split("|") if p and p != path]
    recent.insert(0, path)
    cp["main"]["recent"] = "|".join(recent[:8])
    cp["main"]["last_form"] = form
    save_ini(cp)
    return jsonify({"ok": True, "data": data})


@app.route("/api/browse", methods=["POST"])
def api_browse():
    mode = (request.json or {}).get("mode", "open")
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        ft = [("엑셀 파일", "*.xlsm *.xlsx"), ("모든 파일", "*.*")]
        if mode == "save":
            p = filedialog.asksaveasfilename(defaultextension=".xlsm", filetypes=ft)
        else:
            p = filedialog.askopenfilename(filetypes=ft)
        root.destroy()
        return jsonify({"ok": True, "path": p or ""})
    except Exception as e:
        return jsonify({"ok": False,
                        "msg": "파일 선택 창을 열 수 없습니다(%s). 경로를 직접 입력하십시오." % e})


@app.route("/api/list", methods=["POST"])
def api_list():
    """폴더 안의 엑셀 파일과 하위 폴더를 돌려준다."""
    d = (request.json or {}).get("dir", "").strip().strip('"')
    if not d:
        cp = load_ini()
        recent = [p for p in cp["main"].get("recent", "").split("|") if p]
        d = os.path.dirname(recent[0]) if recent else os.getcwd()
    if not os.path.isdir(d):
        return jsonify({"ok": False, "msg": "폴더가 아닙니다: %s" % d})
    dirs, files = [], []
    try:
        for n in sorted(os.listdir(d)):
            if n.startswith("."):
                continue
            full = os.path.join(d, n)
            if os.path.isdir(full):
                dirs.append(n)
            elif os.path.splitext(n)[1].lower() in (".xlsx", ".xlsm"):
                files.append(n)
    except Exception as e:
        return jsonify({"ok": False, "msg": "폴더를 읽을 수 없습니다: %s" % e})
    parent = os.path.dirname(os.path.abspath(d))
    return jsonify({"ok": True, "dir": os.path.abspath(d),
                    "parent": parent if parent != d else "",
                    "dirs": dirs[:300], "files": files[:300],
                    "sep": os.sep})


@app.route("/api/save", methods=["POST"])
def api_save():
    body = request.json or {}
    target = (body.get("path") or STATE["path"] or "").strip().strip('"')
    if not target:
        return jsonify({"ok": False, "msg": "저장 경로가 없습니다."})
    if body.get("backup") and os.path.exists(target) and target == STATE["path"]:
        try:
            bak = "%s.bak_%s%s" % (os.path.splitext(target)[0],
                                   time.strftime("%Y%m%d_%H%M%S"),
                                   os.path.splitext(target)[1])
            shutil.copy2(target, bak)
        except Exception:
            pass
    try:
        write_workbook(body.get("rows", []), body.get("params", {}), target)
    except Exception as e:
        return jsonify({"ok": False, "msg": "저장 실패: %s" % e})
    return jsonify({"ok": True, "path": target, "name": os.path.basename(target)})


@app.route("/api/alive", methods=["POST", "GET"])
def api_alive():
    LAST_PING[0] = time.time()
    CONNECTED[0] = True
    return jsonify({"ok": True})


@app.route("/api/shutdown", methods=["POST"])
def api_shutdown():
    threading.Timer(0.4, lambda: os._exit(0)).start()
    return jsonify({"ok": True})


def watchdog():
    started = time.time()
    while True:
        time.sleep(5)
        if CONNECTED[0]:
            if time.time() - LAST_PING[0] > 20:
                os._exit(0)
        elif time.time() - started > 300:
            os._exit(0)


# ---------------------------------------------------------------- 화면

HTML = r"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<title>SW개발비 FP 산정양식 편집기</title>
<style>
*{box-sizing:border-box}
body{margin:0;font-family:"맑은 고딕","Malgun Gothic",sans-serif;font-size:13px;color:#222;background:#f4f5f7}
#bar{display:flex;align-items:center;gap:6px;padding:8px 10px;background:#2c3e50;color:#fff}
#bar b{font-size:14px;margin-right:8px}
#fname{flex:1;color:#cfd8dc;font-size:12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
button{font-family:inherit;font-size:12px;padding:5px 10px;border:1px solid #bbb;background:#fff;border-radius:3px;cursor:pointer}
button:hover{background:#eef3f8}
#tabs{display:flex;gap:2px;padding:8px 10px 0}
.tab{padding:7px 16px;border:1px solid #ccc;border-bottom:none;background:#e4e7ea;border-radius:4px 4px 0 0;cursor:pointer}
.tab.on{background:#fff;font-weight:700}
.pane{display:none;background:#fff;border:1px solid #ccc;margin:0 10px 10px;padding:10px}
.pane.on{display:block}
.toolbar{display:flex;align-items:center;gap:6px;margin-bottom:8px;flex-wrap:wrap}
.toolbar input[type=text]{padding:4px 6px;border:1px solid #bbb;border-radius:3px}
#wrap{max-height:calc(100vh - 265px);overflow:auto;border:1px solid #ddd}
table{border-collapse:collapse;width:100%;table-layout:fixed}
th,td{border:1px solid #d7d7d7;padding:0;font-size:12px}
th{background:#eceff1;position:sticky;top:0;z-index:2;padding:5px 3px;font-weight:700}
td input,td select{width:100%;border:none;padding:4px 5px;font-family:inherit;font-size:12px;background:transparent}
td input:focus,td select:focus{outline:2px solid #4a90d9;background:#fffde7}
td.ro{background:#f7f7f7;text-align:center;color:#555}
td.num input{text-align:right}
tr.bad td.ro{background:#fdecea;color:#c0392b;font-weight:700}
.idx{text-align:center;color:#888;background:#f7f7f7}
#sum{margin-top:8px;padding:8px;background:#f0f4f8;border:1px solid #dde3e8;font-size:12px}
#sum table{width:auto;table-layout:auto}
#sum th,#sum td{padding:4px 10px;text-align:right;background:#fff}
#sum th{position:static;background:#eceff1;text-align:center}
.form{max-width:1050px}
.form h3{font-size:13px;margin:16px 0 6px;padding-left:6px;border-left:4px solid #2c3e50}
.form table{table-layout:auto}
.form th{position:static;text-align:left;width:180px;padding:6px 8px;vertical-align:top}
.form td{padding:4px 6px}
.form select,.form input{width:100%;padding:5px;border:1px solid #bbb;border-radius:3px;font-family:inherit;font-size:12px}
.right{text-align:right}
.big{font-size:15px;font-weight:700;color:#12507b}
.warn{color:#c0392b;font-weight:700}
.msg{position:fixed;right:16px;bottom:16px;background:#333;color:#fff;padding:9px 14px;border-radius:4px;opacity:0;transition:.25s;z-index:99}
.msg.on{opacity:.95}
#open{padding:24px;max-width:900px}
#open input[type=text]{width:100%;padding:6px;border:1px solid #bbb}
#open label{display:block;padding:8px 10px;border:1px solid #ccc;border-radius:4px;margin:5px 0;cursor:pointer;background:#fff}
#open label.on{background:#eaf3fb;border-color:#4a90d9}
#open label.off{color:#999;background:#f7f7f7;border-color:#e0e0e0;cursor:default}
#open .grp{margin:14px 0 4px;font-weight:700;color:#2c3e50;font-size:12px;
  border-bottom:1px solid #ccd;padding-bottom:3px}
#open .code{display:inline-block;min-width:52px;color:#12507b;font-weight:700}
#open label.off .code{color:#aaa}
#open .soon{font-size:11px;color:#b06000;background:#fff3e0;border:1px solid #ffcc80;
  border-radius:3px;padding:1px 5px;margin-left:4px}
#open .fn{font-size:11px;color:#999;margin:2px 0 0 52px}
#open h3{font-size:13px;margin:18px 0 6px;padding-left:6px;border-left:4px solid #2c3e50}
.rec{display:block;margin:3px 0;color:#12507b;text-decoration:none;font-size:12px}
small{color:#666}
</style></head><body>

<div id="bar">
  <b>SW개발비 FP 산정양식 편집기</b>
  <button onclick="openDlg()">파일 열기</button>
  <button onclick="save(false)">저장</button>
  <button onclick="save(true)">다른 이름으로 저장</button>
  <span id="fname">파일이 열려 있지 않습니다.</span>
  <span id="dirty" style="color:#ffcc80;font-size:12px"></span>
  <button onclick="quit()">종료</button>
</div>

<div id="open">
  <h3>1. 산정양식 파일</h3>
  <p><input type="text" id="path" placeholder="파일 경로 (직접 입력하거나 아래에서 고르십시오)"></p>
  <p><button onclick="browse()">찾아보기</button>
     <button onclick="listDir('')">폴더에서 고르기</button></p>
  <div id="browser" style="display:none">
    <p style="margin:4px 0"><input type="text" id="dir" style="width:75%"
       onchange="listDir(this.value)"> <button onclick="listDir(document.getElementById('dir').value)">이동</button></p>
    <div id="entries" style="max-height:240px;overflow:auto;border:1px solid #ccc;padding:6px;background:#fff"></div>
  </div>
  <h3>2. 양식 선택</h3>
  <div id="forms"></div>
  <p><button onclick="doOpen()" style="padding:8px 22px;font-size:13px">열기</button>
     <span id="guessMsg"></span></p>
  <div id="recent"></div>
</div>

<div id="main" style="display:none">
<div id="tabs">
  <div class="tab on" data-p="p1" onclick="tab('p1')">① FP 산정</div>
  <div class="tab" data-p="p2" onclick="tab('p2')">② 개발비 산정</div>
  <div class="tab" data-p="p3" onclick="tab('p3')">③ 검증</div>
</div>

<div class="pane on" id="p1">
  <div class="toolbar">
    <button onclick="addRows(1)">행 추가</button>
    <button onclick="addRows(10)">10행 추가</button>
    <button onclick="delRows()">선택행 삭제</button>
    <button onclick="dupRow()">선택행 복사</button>
    <span id="methodBox"></span>
    <span style="margin-left:10px">검색 <input type="text" id="q" oninput="render()" placeholder="기능명·설명"></span>
    <label style="padding:0"><input type="checkbox" id="onlyBad" onchange="render()" style="width:auto"> 오류행만</label>
    <span id="cnt" style="margin-left:auto"></span>
  </div>
  <div id="wrap"><table id="tb"></table></div>
  <div id="sum"></div>
</div>

<div class="pane" id="p2"><div class="form" id="f2"></div></div>
<div class="pane" id="p3"><div id="v3"></div></div>
</div>

<div class="msg" id="msg"></div>

<script>
var D=null, ROWS=[], COLS=[], COST={}, PARAMS={}, TYPES=[], DEVS=[], SW={},
    ADJT={}, ADJL={}, PATH='', FORM='', FORMS=[], VIEW=[], DIRTY=false;

function mark(){DIRTY=true;var e=document.getElementById('dirty');if(e)e.textContent='● 저장 안 됨';}
function clean(){DIRTY=false;var e=document.getElementById('dirty');if(e)e.textContent='';}
function toast(t){var m=document.getElementById('msg');m.textContent=t;m.classList.add('on');
  clearTimeout(m._t);m._t=setTimeout(function(){m.classList.remove('on')},2200);}
function num(v){if(v===''||v===null||v===undefined)return null;var n=parseFloat(v);return isNaN(n)?null:n;}
function won(v){if(v===null||v===undefined||isNaN(v))return '-';return Math.round(v).toLocaleString('ko-KR');}
function fx(v,d){if(v===null||v===undefined||isNaN(v))return '-';return v.toFixed(d);}
function esc(s){return (s===null||s===undefined)?'':String(s)
  .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/"/g,'&quot;');}
function has(k){for(var i=0;i<COLS.length;i++){if(COLS[i].k===k)return true;}return false;}

/* ---- 원본 양식 VBA(FPR/FPV) 와 동일 ---- */
function fpRating(t,ftr,det){
  var r=null;
  if(t==='EO'||t==='EQ'){
    if(det===0&&ftr===0) r=0;
    if(det>0&&det<20&&ftr<2) r='L';
    if(det>19&&ftr<2) r='A';
    if((ftr===2||ftr===3)&&det<=5) r='L';
    if((ftr===2||ftr===3)&&det>5&&det<20) r='A';
    if((ftr===2||ftr===3)&&det>19) r='H';
    if(ftr>3&&det<6) r='A';
    if(ftr>3&&det>=6) r='H';
  }else if(t==='EI'){
    if(det===0&&ftr===0) r=0;
    if(det>0&&det<16&&ftr<2) r='L';
    if(det>15&&ftr<2) r='A';
    if(ftr===2&&det<=4) r='L';
    if(ftr===2&&det>4&&det<16) r='A';
    if(ftr===2&&det>=16) r='H';
    if(ftr>2&&det<5) r='A';
    if(ftr>2&&det>=5) r='H';
  }else if(t==='ILF'||t==='EIF'){
    if(det===0&&ftr===0) r=0;
    if(det>0&&det<51&&ftr<2) r='L';
    if(det>50&&ftr<2) r='A';
    if(ftr>=2&&ftr<=5&&det<20) r='L';
    if(ftr>=2&&ftr<=5&&det>19&&det<51) r='A';
    if(ftr>=2&&ftr<=5&&det>=51) r='H';
    if(ftr>=6&&det<=19) r='A';
    if(ftr>=6&&det>19) r='H';
  }
  return r;
}
var FPV={EI:{L:3,A:4,H:6},EQ:{L:3,A:4,H:6},EO:{L:4,A:5,H:7},
         ILF:{L:7,A:10,H:15},EIF:{L:5,A:7,H:10}};

function calcRow(r){
  var t=(r.type||'').toUpperCase();
  r._cx=''; r._wt=null; r._wtadj=null; r._err='';
  if(!t){ if(r.app||r.proc||r.desc) r._err='FP유형 미입력'; return r;}
  if(PARAMS.method==='간이법'){
    r._wt=(SW[t]===undefined)?null:SW[t];
    if(r._wt===null) r._err='FP유형 값 확인';
    else if(!r.proc) r._err='단위프로세스명 미입력';
    return r;
  }
  var ftr=num(r.ftr), det=num(r.det);
  if(ftr===null||det===null){ r._err='RET/FTR 또는 DET 미입력'; return r;}
  var g=fpRating(t,ftr,det);
  if(g===null){ r._err='복잡도 판정 불가 (RET/FTR·DET 조합 확인)'; return r;}
  if(g===0){ r._cx='0'; r._err='RET/FTR·DET 가 0이어서 기능점수 0'; return r;}
  r._cx=g; r._wt=FPV[t][g];
  if(!r.proc) r._err='단위프로세스명 미입력';
  return r;
}

/* ---- 재개발 변경률 (2016 양식) : 원본 시트 E40~E53, M열 수식과 동일 ---- */
function rwNum(k){var x=(COST.rework||{})[k];if(!x)return null;
  var v=String(x.value||'');return (v===''||v==='(선택)')?null:num(v);}
function reworkChain(){
  var R={ok:false,diff:null,level:null,effort:null,famil:null,integ:null,
         dataSize:0,tranSize:0,size:0,dataRate:0,tranRate:0};
  if(COST.kind!=='kosa2016') return R;
  var d1=rwNum('d1'),d2=rwNum('d2'),d3=rwNum('d3');
  R.level=(d1===null||d2===null||d3===null)?null:(d1+d2+d3)/3;
  R.effort=rwNum('effort'); R.famil=rwNum('famil');
  var code=rwNum('coderate')||0, test=rwNum('testrate')||0;
  var dNum=0,tNum=0;
  ROWS.forEach(function(r){
    if(r.dev!=='수정 후 재사용') return;
    var g=num(r.chg); if(g===null||r._wt===null) return;
    var t=(r.type||'').toUpperCase();
    if(t==='ILF'||t==='EIF'){R.dataSize+=r._wt; dNum+=g*r._wt;}
    else if(t==='EI'||t==='EO'||t==='EQ'){R.tranSize+=r._wt; tNum+=g*r._wt;}
  });
  R.size=R.dataSize+R.tranSize;
  R.dataRate=R.dataSize>0?dNum/R.dataSize:0;
  R.tranRate=R.tranSize>0?tNum/R.tranSize:0;
  if(R.dataRate===0&&R.tranRate===0&&R.size===0) return R;
  if(R.dataRate!==0&&R.tranRate!==0)
    R.integ=(R.dataSize/R.size*R.dataRate)+(R.tranSize/R.size*R.tranRate);
  else if(R.dataRate===0) R.integ=R.tranSize/R.size*R.tranRate;
  else R.integ=R.dataSize/R.size*R.dataRate;
  R.diff=0.4*R.integ+0.3*code+0.3*test;
  R.ok=(R.effort!==null&&R.level!==null&&R.famil!==null);
  return R;
}
function applyRework(R){
  ROWS.forEach(function(r){r._wtadj=null;});
  if(!R.ok||R.diff===null) return;
  ROWS.forEach(function(r){
    if(r.dev!=='수정 후 재사용') return;
    if(num(r.chg)===null||r._wt===null) return;
    r._wtadj=(R.diff<=0.5)
      ? r._wt*(R.effort+R.diff*100*(1+0.02*(R.level*R.famil)))/100
      : r._wt*(R.effort+R.diff*100+(R.level*R.famil))/100;
  });
}
function calcAll(){ROWS.forEach(calcRow);var R=reworkChain();applyRework(R);return R;}

/* ---- 표 ---- */
function render(){
  var R=calcAll();
  var q=document.getElementById('q').value.trim().toLowerCase();
  var only=document.getElementById('onlyBad').checked;
  VIEW=[];
  for(var i=0;i<ROWS.length;i++){
    var r=ROWS[i];
    if(only&&!r._err) continue;
    if(q){var s=((r.app||'')+(r.biz||'')+(r.proc||'')+(r.desc||'')+(r.remark||'')).toLowerCase();
      if(s.indexOf(q)<0) continue;}
    VIEW.push(i);
  }
  var h=['<colgroup><col width="34"><col width="38">'];
  COLS.forEach(function(c){h.push('<col width="'+c.w+'">');});
  h.push('</colgroup><tr><th></th><th>#</th>');
  COLS.forEach(function(c){h.push('<th>'+esc(c.label)+'</th>');});
  h.push('</tr>');
  for(var k=0;k<VIEW.length;k++){
    var i=VIEW[k], r=ROWS[i];
    h.push('<tr class="'+(r._err?'bad':'')+'" data-i="'+i+'">',
      '<td style="text-align:center"><input type="checkbox" class="ck" style="width:auto"></td>',
      '<td class="idx">'+(i+1)+'</td>');
    for(var m=0;m<COLS.length;m++){
      var c=COLS[m];
      if(c.t==='ro'){
        var v='';
        if(c.k==='cx') v=(r._cx===''?(r._err?'!':''):r._cx);
        else if(c.k==='wt') v=(r._wt===null?'':r._wt);
        else if(c.k==='wtadj') v=(r._wtadj===null?'':r._wtadj.toFixed(2));
        h.push('<td class="ro" title="'+esc(r._err)+'">'+v+'</td>');
      }else if(c.t==='sel'){
        var list=(c.k==='type')?TYPES:DEVS, o='<option value=""></option>';
        for(var z=0;z<list.length;z++){
          o+='<option'+((r[c.k]||'')===list[z]?' selected':'')+'>'+esc(list[z])+'</option>';}
        h.push('<td><select onchange="up('+i+',\''+c.k+'\',this.value)">'+o+'</select></td>');
      }else{
        h.push('<td class="'+(c.t==='num'?'num':'')+'"><input value="'+esc(r[c.k])+
               '" onchange="up('+i+',\''+c.k+'\',this.value)"></td>');
      }
    }
    h.push('</tr>');
  }
  document.getElementById('tb').innerHTML=h.join('');
  document.getElementById('cnt').textContent='표시 '+VIEW.length+' / 전체 '+ROWS.length+'행';
  summary(R); form(R); verify(R);
}
function up(i,k,v){ROWS[i][k]=v;mark();
  if(k==='type'||k==='ftr'||k==='det'||k==='dev'||k==='chg') render();
  else {var R=calcAll();summary(R);form(R);}}

function agg(){
  var a={},tot={n:0,fp:0,adj:0},dev={};
  TYPES.forEach(function(t){a[t]={n:0,fp:0,adj:0};});
  DEVS.forEach(function(d){dev[d]={n:0,fp:0};});
  ROWS.forEach(function(r){
    var t=(r.type||'').toUpperCase(); if(!a[t])return;
    a[t].n++; tot.n++;
    if(r._wt){a[t].fp+=r._wt; tot.fp+=r._wt;}
    var c;
    if(COST.kind==='kosa2016'){
      if(r.dev==='수정 후 재사용') c=r._wtadj||0;
      else if(r.dev==='수정 없이 재사용') c=0;
      else c=r._wt||0;
    }else{ c=r._wt||0; }
    a[t].adj+=c; tot.adj+=c;
    if(r.dev&&dev[r.dev]){dev[r.dev].n++; dev[r.dev].fp+=(r._wt||0);}
  });
  return {a:a,tot:tot,dev:dev};
}
function targetFP(){var g=agg();return (COST.kind==='kosa2016')?g.tot.adj:g.tot.fp;}

function summary(){
  var g=agg(),h=['<table><tr><th>구분</th>'];
  TYPES.forEach(function(t){h.push('<th>'+t+'</th>');});
  h.push('<th>계</th></tr><tr><th>기능수</th>');
  TYPES.forEach(function(t){h.push('<td>'+g.a[t].n+'</td>');});
  h.push('<td class="big">'+g.tot.n+'</td></tr><tr><th>기능점수</th>');
  TYPES.forEach(function(t){h.push('<td>'+fx(g.a[t].fp,1)+'</td>');});
  h.push('<td class="big">'+fx(g.tot.fp,1)+'</td></tr>');
  if(COST.kind==='kosa2016'){
    h.push('<tr><th>개발대상 FP</th>');
    TYPES.forEach(function(t){h.push('<td>'+fx(g.a[t].adj,1)+'</td>');});
    h.push('<td class="big">'+fx(g.tot.adj,1)+'</td></tr>');
  }
  h.push('<tr><th>비중</th>');
  TYPES.forEach(function(t){h.push('<td>'+(g.tot.fp>0?(g.a[t].fp/g.tot.fp*100).toFixed(1)+'%':'-')+'</td>');});
  h.push('<td>'+(g.tot.fp>0?'100.0%':'-')+'</td></tr></table>');
  if(COST.kind==='kosa2016'){
    h.push('<div style="margin-top:6px">개발유형별 : ');
    DEVS.forEach(function(d){h.push(esc(d)+' '+g.dev[d].n+'건('+fx(g.dev[d].fp,1)+'FP) &nbsp; ');});
    h.push('</div>');
  }
  document.getElementById('sum').innerHTML=h.join('');
}

/* ---- 개발비 산정 ---- */
function sizeAdj2024(fp){
  if(fp<500) return 1.28;
  if(fp>3000) return 1.153;
  return 0.4057*Math.pow(Math.log(fp)-7.1978,2)+0.8878;
}
function sizeAdj2016(fp){return fp<300?0.65:0.108*Math.log(fp)+0.2229;}
function adjVal(kind,label){
  var t=ADJT[kind];
  for(var i=0;i<t.length;i++){if(t[i][0]===label)return t[i][1];}
  return t[0][1];
}
function cost2024(){
  var fp=targetFP(), s=sizeAdj2024(fp);
  var v={link:adjVal('link',COST.adj.link),perf:adjVal('perf',COST.adj.perf),
         env:adjVal('env',COST.adj.env),sec:adjVal('sec',COST.adj.sec)};
  var price=num(COST.unitPrice)||0;
  var base=fp*price*s*v.link*v.perf*v.env*v.sec;
  var profit=base*(num(COST.profitRate)||0);
  var exp=0;(COST.expenses||[]).forEach(function(e){exp+=num(e.amount)||0;});
  return {fp:fp,size:s,v:v,price:price,base:base,profit:profit,exp:exp,total:base+profit+exp};
}
function cost2016(){
  var fp=targetFP(), appSum=0,appCo=0,langSum=0,langCo=0,q=0;
  (COST.apps||[]).forEach(function(x){var s=num(x.share)||0;appSum+=s;appCo+=s*x.coef;});
  (COST.langs||[]).forEach(function(x){var s=num(x.share)||0;langSum+=s;langCo+=s*x.coef;});
  (COST.quality||[]).forEach(function(x){q+=num(x.val)||0;});
  var size=sizeAdj2016(fp), qc=q*0.025+1, price=num(COST.unitPrice)||0;
  var stages=[],base=0;
  (COST.stages||[]).forEach(function(st,i){
    var unit=(num(st.w)||0)*price, useLang=(i>=2);
    var amt=fp*unit*appCo*size*qc*(useLang?langCo:1);
    stages.push({name:st.name,w:st.w,unit:unit,amt:amt,lang:useLang});
    base+=amt;
  });
  var profit=base*(num(COST.profitRate)||0);
  return {fp:fp,appSum:appSum,appCo:appCo,langSum:langSum,langCo:langCo,
          size:size,q:q,qc:qc,price:price,stages:stages,base:base,
          profit:profit,total:base+profit};
}
function selAdj(kind){
  var h='<select onchange="upAdj(\''+kind+'\',this.value)">';
  ADJT[kind].forEach(function(o){
    h+='<option'+(COST.adj[kind]===o[0]?' selected':'')+' value="'+esc(o[0])+'">'+esc(o[0])+'  ('+o[1]+')</option>';});
  return h+'</select>';
}
function form(R){
  var el=document.getElementById('f2');
  if(COST.kind==='none'){
    el.innerHTML='<p>이 파일에서는 개발비 산정 시트를 찾지 못했습니다. FP 산정과 검증만 사용할 수 있습니다.</p>';
    return;
  }
  el.innerHTML=(COST.kind==='kosa2024')?form2024():form2016(R);
}
function form2024(){
  var c=cost2024();
  var h=['<h3>사업 정보 (2024 대가산정 가이드)</h3><table>',
   '<tr><th>사업명(제목)</th><td><input value="'+esc(COST.title)+'" onchange="upC(\'title\',this.value)"></td></tr>',
   '<tr><th>기능점수당 단가(원)</th><td><input value="'+esc(COST.unitPrice)+'" onchange="upC(\'unitPrice\',this.value)"></td></tr>',
   '<tr><th>이윤율</th><td><input value="'+esc(COST.profitRate)+'" onchange="upC(\'profitRate\',this.value)"> <small>0.2 = 20%</small></td></tr></table>',
   '<h3>보정계수</h3><table>',
   '<tr><th>SW규모</th><td>총기능점수 '+fx(c.fp,1)+' FP → <b>'+fx(c.size,4)+'</b><br>',
   '<small>= 0.4057 × (ln FP − 7.1978)² + 0.8878, 500FP 미만 1.28 / 3,000FP 초과 1.153</small></td></tr>'];
  ['link','perf','env','sec'].forEach(function(k){
    h.push('<tr><th>'+ADJL[k]+'</th><td>'+selAdj(k)+'</td></tr>');});
  h.push('</table><h3>직접경비</h3><table><tr><th style="width:150px">구분</th><th>산출내역</th><th style="width:150px">금액(원)</th></tr>');
  (COST.expenses||[]).forEach(function(e,i){
    h.push('<tr><td><input value="'+esc(e.name)+'" onchange="upE('+i+',\'name\',this.value)"></td>',
      '<td><input value="'+esc(e.detail)+'" onchange="upE('+i+',\'detail\',this.value)"></td>',
      '<td><input class="right" value="'+esc(e.amount)+'" onchange="upE('+i+',\'amount\',this.value)"></td></tr>');});
  h.push('<tr><th>합 계</th><td></td><td class="right big">'+won(c.exp)+'</td></tr></table>',
   '<small>원본 양식의 직접경비 입력 행이 2행이므로 2건까지 저장됩니다.</small>',
   '<h3>개발원가 산정</h3><table>',
   '<tr><th>총기능점수</th><td class="right">'+fx(c.fp,1)+' FP</td></tr>',
   '<tr><th>기능점수당 단가</th><td class="right">'+won(c.price)+' 원</td></tr>',
   '<tr><th>보정계수 곱</th><td class="right">'+fx(c.size*c.v.link*c.v.perf*c.v.env*c.v.sec,4)+
     ' <small>(규모 '+fx(c.size,4)+' × 연계 '+c.v.link+' × 성능 '+c.v.perf+' × 운영환경 '+c.v.env+' × 보안 '+c.v.sec+')</small></td></tr>',
   '<tr><th>보정 후 개발원가</th><td class="right big">'+won(c.base)+' 원</td></tr>',
   '<tr><th>이윤</th><td class="right">'+won(c.profit)+' 원</td></tr>',
   '<tr><th>직접경비</th><td class="right">'+won(c.exp)+' 원</td></tr>',
   '<tr><th>소프트웨어 개발비(부가세 별도)</th><td class="right big">'+won(c.total)+' 원</td></tr></table>');
  return h.join('');
}
function shareRows(grp,key,items){
  var h='';
  items.forEach(function(x,i){
    h+='<div style="display:flex;gap:6px;align-items:center;margin:2px 0">'+
       '<span style="flex:1">'+esc(x.label)+'</span>'+
       '<input style="width:90px" class="right" value="'+esc(x[key])+
       '" onchange="upList(\''+grp+'\','+i+',\''+key+'\',this.value)"></div>';});
  return h;
}
function form2016(R){
  var c=cost2016();
  var h=['<h3>사업 정보 (2016 대가산정 가이드 · 상세법)</h3><table>',
   '<tr><th>사업명(제목)</th><td><input value="'+esc(COST.title)+'" onchange="upC(\'title\',this.value)"></td></tr>',
   '<tr><th>기능점수당 단가(원)</th><td><input value="'+esc(COST.unitPrice)+'" onchange="upC(\'unitPrice\',this.value)"></td></tr>',
   '<tr><th>이윤율</th><td><input value="'+esc(COST.profitRate)+'" onchange="upC(\'profitRate\',this.value)"> <small>0.2 = 20%</small></td></tr></table>',
   '<h3>보정계수</h3><table>',
   '<tr><th>어플리케이션 유형</th><td>'+shareRows('apps','share',COST.apps||[]),
   '<div class="'+(Math.abs(c.appSum-1)>1e-9?'warn':'')+'">비중 합 '+fx(c.appSum*100,1)+'% → 보정계수 <b>'+fx(c.appCo,4)+'</b>'+
     (Math.abs(c.appSum-1)>1e-9?' (합이 100%가 아닙니다)':'')+'</div></td></tr>',
   '<tr><th>언어</th><td>'+shareRows('langs','share',COST.langs||[]),
   '<div class="'+(Math.abs(c.langSum-1)>1e-9?'warn':'')+'">비중 합 '+fx(c.langSum*100,1)+'% → 보정계수 <b>'+fx(c.langCo,4)+'</b>'+
     (Math.abs(c.langSum-1)>1e-9?' (합이 100%가 아닙니다)':'')+'</div>'+
     '<small>언어 보정은 구현·시험 단계에만 적용됩니다.</small></td></tr>',
   '<tr><th>규모</th><td>개발대상 기능점수 '+fx(c.fp,1)+' FP → <b>'+fx(c.size,4)+'</b><br>'+
     '<small>= 0.108 × ln(FP) + 0.2229, 300FP 미만 0.65</small></td></tr>',
   '<tr><th>품질 및 특성</th><td>'+shareRows('quality','val',COST.quality||[]),
   '<div>총 영향도 '+fx(c.q,1)+' → 보정계수 <b>'+fx(c.qc,4)+'</b> <small>(= 0.025 × 총영향도 + 1)</small></div></td></tr></table>',
   '<h3>재개발 특성 (수정 후 재사용)</h3><table>'];
  var rwk=COST.rework||{};
  ['effort','d1','d2','d3','famil','coderate','testrate'].forEach(function(k){
    var it=rwk[k]; if(!it) return;
    var inp;
    if(it.options){
      inp='<select onchange="upRw(\''+k+'\',this.value)">';
      it.options.forEach(function(o){inp+='<option'+(String(it.value)===o?' selected':'')+'>'+o+'</option>';});
      inp+='</select>';
    }else{
      inp='<input value="'+esc(it.value)+'" onchange="upRw(\''+k+'\',this.value)">';
    }
    h.push('<tr><th>'+esc(it.label)+'</th><td>'+inp+'</td></tr>');
  });
  h.push('<tr><th>재사용 난이도(평균)</th><td>'+(R.level===null?'(입력 필요)':fx(R.level,2))+'</td></tr>',
   '<tr><th>수정 대상 SW 규모</th><td>데이터 '+fx(R.dataSize,1)+' + 트랜잭션 '+fx(R.tranSize,1)+' = <b>'+fx(R.size,1)+'</b> FP</td></tr>',
   '<tr><th>설계 변경률</th><td>데이터 '+fx(R.dataRate,4)+' / 트랜잭션 '+fx(R.tranRate,4)+
     ' → 통합 '+(R.integ===null?'-':fx(R.integ,4))+'</td></tr>',
   '<tr><th>총 변경률</th><td>'+(R.diff===null?'-':fx(R.diff,4))+
     (R.size>0&&!R.ok?' <span class="warn">재사용 평가노력·난이도·친숙도를 모두 입력해야 변경률 적용 가중치가 계산됩니다.</span>':'')+
     '</td></tr></table>',
   '<h3>개발원가 산정</h3><table><tr><th>단계</th><th class="right">가중치</th><th class="right">단계별 단가</th><th class="right">언어보정</th><th class="right">금액(원)</th></tr>');
  c.stages.forEach(function(s){
    h.push('<tr><td>'+esc(s.name)+'</td><td class="right">'+s.w+'</td><td class="right">'+won(s.unit)+
      '</td><td class="right">'+(s.lang?fx(c.langCo,4):'-')+'</td><td class="right">'+won(s.amt)+'</td></tr>');});
  h.push('<tr><th>합계(보정 후 개발원가)</th><td colspan="3"></td><td class="right big">'+won(c.base)+'</td></tr>',
   '<tr><th>이윤</th><td colspan="3"></td><td class="right">'+won(c.profit)+'</td></tr>',
   '<tr><th>소프트웨어 개발비(부가세 별도)</th><td colspan="3"></td><td class="right big">'+won(c.total)+'</td></tr></table>',
   '<small>이 양식의 개발비 시트에는 직접경비 입력란이 없어 다루지 않습니다.</small>');
  return h.join('');
}
function upC(k,v){COST[k]=v;mark();var R=calcAll();form(R);}
function upAdj(k,v){COST.adj[k]=v;mark();var R=calcAll();form(R);}
function upE(i,k,v){COST.expenses[i][k]=v;mark();var R=calcAll();form(R);}
function upList(grp,i,k,v){COST[grp][i][k]=v;mark();var R=calcAll();form(R);}
function upRw(k,v){COST.rework[k].value=v;mark();render();}
function upMethod(v){PARAMS.method=v;mark();render();}

/* ---- 검증 ---- */
function verify(){
  var issues=[],seen={};
  ROWS.forEach(function(r,i){
    if(r._err) issues.push([i+1,r.proc||r.app||'','오류',r._err]);
    var key=(r.app||'')+'|'+(r.biz||'')+'|'+(r.proc||'');
    if(r.proc){
      if(seen[key]!==undefined) issues.push([i+1,r.proc,'중복','같은 업무 안에 동일한 단위프로세스명이 '+(seen[key]+1)+'행에도 있습니다.']);
      else seen[key]=i;}
    if(r.type&&!r.app) issues.push([i+1,r.proc||'','누락','애플리케이션명이 비어 있습니다.']);
    if(has('dev')){
      if(r.type&&!r.dev) issues.push([i+1,r.proc||'','누락','개발유형이 비어 있습니다.']);
      if(r.dev==='수정 후 재사용'&&num(r.chg)===null)
        issues.push([i+1,r.proc||'','누락','수정 후 재사용인데 설계변경률이 없습니다.']);
    }
    if(r.file_cx&&r._cx&&String(r.file_cx)!==String(r._cx))
      issues.push([i+1,r.proc||'','대조','원본 복잡도 '+r.file_cx+' ↔ 재계산 '+r._cx]);
    if(r.file_wt&&r._wt!==null&&Math.abs(Number(r.file_wt)-Number(r._wt))>1e-9)
      issues.push([i+1,r.proc||'','대조','원본 가중치 '+r.file_wt+' ↔ 재계산 '+r._wt]);
    if(r.file_wtadj&&r._wtadj!==null&&Math.abs(Number(r.file_wtadj)-r._wtadj)>1e-6)
      issues.push([i+1,r.proc||'','대조','원본 변경률 가중치 '+Number(r.file_wtadj).toFixed(3)+' ↔ 재계산 '+r._wtadj.toFixed(3)]);
  });
  var h=['<p>총 '+ROWS.length+'행 중 확인이 필요한 항목 <b>'+issues.length+'</b>건'+
    ' &nbsp;|&nbsp; 양식 '+esc(D.formName)+
    ' &nbsp;|&nbsp; 시트 '+esc(D.sheet)+' (머리글 '+D.headerRow+'행, 데이터 '+D.startRow+'행부터, 산정방법 '+esc(PARAMS.method)+')</p>'];
  if(issues.length){
    h.push('<table><tr><th style="width:60px">행</th><th style="width:220px">단위프로세스명</th><th style="width:70px">구분</th><th>내용</th></tr>');
    issues.slice(0,500).forEach(function(x){
      h.push('<tr><td class="right">'+x[0]+'</td><td>'+esc(x[1])+'</td><td>'+x[2]+'</td><td>'+esc(x[3])+'</td></tr>');});
    h.push('</table>');
    if(issues.length>500) h.push('<p><small>앞의 500건만 표시했습니다.</small></p>');
  }
  h.push('<p><small>"대조"는 원본 파일에 저장돼 있던 계산 결과와 이 프로그램의 재계산 결과가 다른 경우입니다.</small></p>');
  document.getElementById('v3').innerHTML=h.join('');
}

/* ---- 행 편집 ---- */
function blank(){var o={};COLS.forEach(function(c){if(c.t!=='ro')o[c.k]='';});return o;}
function addRows(n){for(var i=0;i<n;i++)ROWS.push(blank());mark();render();
  var w=document.getElementById('wrap');w.scrollTop=w.scrollHeight;}
function checked(){var out=[];
  document.querySelectorAll('#tb tr[data-i]').forEach(function(tr){
    if(tr.querySelector('.ck').checked) out.push(parseInt(tr.dataset.i,10));});
  return out;}
function delRows(){var s=checked();if(!s.length){toast('삭제할 행을 선택하십시오.');return;}
  if(!confirm(s.length+'행을 삭제합니다.'))return;
  s.sort(function(a,b){return b-a;}).forEach(function(i){ROWS.splice(i,1);});mark();render();}
function dupRow(){var s=checked();if(!s.length){toast('복사할 행을 선택하십시오.');return;}
  s.sort(function(a,b){return b-a;}).forEach(function(i){
    ROWS.splice(i+1,0,JSON.parse(JSON.stringify(ROWS[i])));});mark();render();}

/* ---- 파일 ---- */
function tab(p){document.querySelectorAll('.tab').forEach(function(t){t.classList.toggle('on',t.dataset.p===p);});
  document.querySelectorAll('.pane').forEach(function(t){t.classList.toggle('on',t.id===p);});}
function openDlg(){document.getElementById('open').style.display='block';
  document.getElementById('main').style.display='none';}
function firstReady(){for(var i=0;i<FORMS.length;i++){if(FORMS[i].ready)return FORMS[i].key;}return '';}
function paintForms(sel){
  var ok=false;
  FORMS.forEach(function(f){if(f.key===sel&&f.ready)ok=true;});
  FORM = ok ? sel : firstReady();
  var h='', g='';
  FORMS.forEach(function(f){
    if(f.group!==g){g=f.group; h+='<div class="grp">'+esc(g)+'</div>';}
    var code=f.code?('<span class="code">'+esc(f.code)+'</span> '):'';
    if(f.ready){
      h+='<label class="'+(f.key===FORM?'on':'')+'" onclick="paintForms(\''+f.key+'\')">'+
         '<input type="radio" name="fm" style="width:auto" '+(f.key===FORM?'checked':'')+'> '+
         code+esc(f.name)+
         (f.file?'<div class="fn">'+esc(f.file)+'</div>':'')+'</label>';
    }else{
      h+='<label class="off" title="아직 지원하지 않는 양식입니다">'+
         '<input type="radio" disabled style="width:auto"> '+code+esc(f.name)+
         ' <span class="soon">추후 개발예정</span>'+
         (f.file?'<div class="fn">'+esc(f.file)+'</div>':'')+'</label>';
    }
  });
  document.getElementById('forms').innerHTML=h;
}
function browse(){fetch('/api/browse',{method:'POST',headers:{'Content-Type':'application/json'},
  body:JSON.stringify({mode:'open'})}).then(function(r){return r.json();}).then(function(j){
    if(j.ok&&j.path){pickPath(j.path);}
    else if(!j.ok){toast(j.msg);listDir('');}});}
function guess(){
  var p=document.getElementById('path').value;
  if(!p)return;
  fetch('/api/guess',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({path:p})}).then(function(r){return r.json();}).then(function(j){
    if(j.ok&&j.form){paintForms(j.form);
      document.getElementById('guessMsg').innerHTML='<small>시트 구성으로 보아 '+j.form+' 양식으로 판단했습니다. 다르면 위에서 바꾸십시오.</small>';}
    else document.getElementById('guessMsg').innerHTML='<small>양식을 판단하지 못했습니다. 직접 선택하십시오.</small>';
  });}
function doOpen(){
  var p=document.getElementById('path').value;
  fetch('/api/open',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({path:p,form:FORM})}).then(function(r){return r.json();}).then(function(j){
    if(!j.ok){alert(j.msg);return;}
    load(j.data);});}
function load(d){
  D=d;ROWS=d.rows;COLS=d.cols;COST=d.cost;PARAMS=d.params;TYPES=d.fpTypes;DEVS=d.devTypes;
  SW=d.simpleWeight;ADJT=d.adjTable;ADJL=d.adjLabel;PATH=d.path;FORM=d.form;
  document.getElementById('fname').textContent='['+d.form+'] '+d.path;
  document.getElementById('open').style.display='none';
  document.getElementById('main').style.display='block';
  var mb=document.getElementById('methodBox');
  if(d.params.methodCell){
    var o='';['상세법','간이법'].forEach(function(x){
      o+='<option'+(PARAMS.method===x?' selected':'')+'>'+x+'</option>';});
    mb.innerHTML=' 산정방법 <select onchange="upMethod(this.value)">'+o+'</select>';
  }else mb.innerHTML='';
  if(!ROWS.length) ROWS.push(blank());
  render();clean();
  toast('불러왔습니다. '+ROWS.length+'행 · '+d.formName);
}
function save(as){
  var body={rows:ROWS,params:{method:PARAMS.method,cost:COST},backup:true};
  if(as){
    fetch('/api/browse',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({mode:'save'})}).then(function(r){return r.json();}).then(function(j){
      var p=(j.ok&&j.path)?j.path:prompt('저장할 파일 경로를 입력하십시오.',PATH);
      if(!p)return; body.path=p; doSave(body);});
  }else{ body.path=PATH;
    if(!confirm('원본 파일에 덮어씁니다. (같은 폴더에 .bak 사본을 만듭니다)'))return; doSave(body);}
}
function doSave(body){
  fetch('/api/save',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify(body)}).then(function(r){return r.json();}).then(function(j){
    if(!j.ok){alert(j.msg);return;}
    PATH=j.path;document.getElementById('fname').textContent='['+FORM+'] '+j.path;
    clean();toast('저장했습니다: '+j.name);});
}
function quit(){
  if(DIRTY&&!confirm('저장하지 않은 변경이 있습니다. 그래도 종료하시겠습니까?'))return;
  if(!DIRTY&&!confirm('프로그램을 종료합니다.'))return;
  DIRTY=false;fetch('/api/shutdown',{method:'POST'});setTimeout(function(){window.close();},300);}

window.onbeforeunload=function(){if(DIRTY)return '저장하지 않은 변경이 있습니다.';};
setInterval(function(){fetch('/api/alive',{method:'POST'});},4000);
document.getElementById('path').addEventListener('change',guess);
fetch('/api/init').then(function(r){return r.json();}).then(function(j){
  FORMS=j.forms; paintForms(j.data?j.data.form:'');
  if(j.data) load(j.data);
  var h='';
  if(j.recent&&j.recent.length){h='<h3>최근 파일</h3>';
    j.recent.forEach(function(p){h+='<a class="rec" href="#" onclick="document.getElementById(\'path\').value=this.textContent;guess();return false;">'+p+'</a>';});}
  document.getElementById('recent').innerHTML=h;
});
</script></body></html>
"""


# ---------------------------------------------------------------- 기동

def arg_path():
    """실행 인자로 준 파일 경로만 본다. 없으면 화면에서 고른다."""
    skip = False
    for a in sys.argv[1:]:
        if skip:
            skip = False
            continue
        if a == "--form":
            skip = True
            continue
        if a.startswith("-"):
            continue
        p = a.strip().strip('"')
        if os.path.exists(p):
            return p
        print("파일을 찾을 수 없습니다:", p)
    return None


def arg_form():
    for i, a in enumerate(sys.argv):
        if a == "--form" and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
        if a.startswith("--form="):
            return a.split("=", 1)[1]
    return None


def open_browser(url):
    chrome = find_chrome()
    proc = None
    if chrome:
        profile = os.path.join(tempfile.gettempdir(), ".chrome_calc_ui")
        try:
            proc = subprocess.Popen(
                [chrome, "--app=" + url, "--window-size=1500,950",
                 "--no-first-run", "--no-default-browser-check",
                 "--user-data-dir=" + profile], **_no_window())
        except Exception:
            proc = None
    if proc is not None:
        # 창이 붙을 때까지 기다린다. 크롬은 실행 프로세스가 먼저 끝나고
        # 창만 남는 경우가 있어 프로세스 종료만으로 판단하지 않는다.
        waited = 0.0
        while waited < 40:
            if CONNECTED[0]:
                break
            if proc.poll() is not None and waited > 12:
                break
            time.sleep(0.5)
            waited += 0.5
        if CONNECTED[0]:
            if proc.poll() is None:
                try:
                    proc.wait()          # 창을 닫으면 프로그램도 종료
                except Exception:
                    pass
                os._exit(0)
            return                       # 창은 살아 있음. 감시 스레드가 종료를 맡는다
    try:
        webbrowser.open(url)             # 크롬이 없거나 창이 뜨지 않은 경우에만
    except Exception:
        pass


def main(only=None):
    setup_output()
    global VISIBLE
    if only:
        VISIBLE = [k for k in FORMS if k in only]
    no_browser = "--no-browser" in sys.argv

    path = arg_path()
    if path:
        form = arg_form() or guess_form(path)
        if form not in VISIBLE or not is_ready(form):
            print("양식을 판단하지 못했습니다. 화면에서 선택하십시오:", path)
        else:
            try:
                read_workbook(path, form)
                print("열었습니다: [%s] %s" % (form, path))
            except Exception as e:
                print("열기 실패:", e)

    port = pick_port()
    url = "http://127.0.0.1:%d/" % port
    cp = load_ini()
    cp["main"]["last_port"] = str(port)
    save_ini(cp)

    print("=" * 60)
    print(APP_NAME, "v" + VERSION)
    print("화면 주소:", url)
    print("창이 뜨지 않으면 브라우저에 위 주소를 직접 입력하십시오.")
    print("=" * 60)

    threading.Thread(target=watchdog, daemon=True).start()
    if not no_browser:
        threading.Thread(target=lambda: (time.sleep(1.0), open_browser(url)),
                         daemon=True).start()

    import logging
    logging.getLogger("werkzeug").setLevel(logging.ERROR)
    try:
        app.run(host="127.0.0.1", port=port, debug=False,
                use_reloader=False, threaded=True)
    except Exception as e:
        print("서버 기동 실패:", e)
        try:
            input("엔터를 누르면 종료합니다.")
        except Exception:
            pass


if __name__ == "__main__":
    main()