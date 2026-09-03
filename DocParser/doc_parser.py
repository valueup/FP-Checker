# -*- coding: utf-8 -*-
"""산출물 읽기 — UI설계서(pptx)와 테이블정의서(xlsx)에서 원자료를 뽑는다.

읽기만 한다. 판정도, 산정도, 대조도 하지 않는다.

이 파일에는 '산출물을 어떻게 읽을 것인가' 규칙만 둔다. 물리 테이블명 형식,
표 머리글 판별, 시트 이름, 논리파일 그룹 키가 여기 있다. 사업이 바뀌면
아래 상수 구역만 고치면 된다. 가중치·복잡도 매트릭스는 Calculator 에 있다.

돌려주는 것
  화면  {"id","name","file","slides","attrs","funcs","tables"}
  테이블 {"name","kor","system","file","columns"}

필요한 것은 openpyxl 하나다. lxml 은 있으면 쓰고 없으면 기본 XML 파서로 돈다.

읽은 결과는 JSON 한 벌로 저장한다. comparator·calculator·reporter 는 산출물을
다시 읽지 않고 이 파일만 읽으면 된다. 형식은 아래 dump() 를 보라.

단독 실행하면 두 폴더를 읽어 요약을 찍고, 경로를 주면 저장한다.
  python -m DocParser.doc_parser <UI설계서폴더> <테이블정의서폴더> [저장경로] [--pretty]

저장경로에 폴더를 주면 doc_parser_YYYYMMDD_HHMM.json 이름을 붙여 그 안에 넣는다.
"""

import datetime
import glob
import json
import os
import re
import shutil
import sys
import time
import zipfile

# 표를 읽는 데 XML 파서가 필요하다. lxml 이 있으면 빠르므로 그쪽을 쓰고,
# 없으면 파이썬에 기본으로 들어 있는 것을 쓴다. 결과는 같다.
try:
    from lxml import etree as _xml
    XML_BACKEND = "lxml"
except ImportError:
    import xml.etree.ElementTree as _xml
    XML_BACKEND = "ElementTree"

# ======================================================================
# 산출물 판독 규칙
# ======================================================================
UI_GLOB = "*UI설계서*.pptx"
TBL_GLOB = "*테이블정의서*.xlsx"

# pptx 표 안의 글자를 담는 네임스페이스
DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"

# 물리 테이블명. 예) COM_CM001_TB
TABLE_NAME = re.compile(r"^[A-Z][A-Z0-9]*_[A-Z0-9]+_TB$")

# 화면ID. 예) UI-A01-001
SCREEN_ID = re.compile(r"^UI-[A-Z]\d{2}-[0-9A-Za-z\-]+$")

# 논리파일 그룹 키. COM_CM001A_TB, COM_CM001B_TB → COM_CM_001
GROUP_KEY = re.compile(r"^([A-Z]+)_([A-Z]{2,})(\d+)([A-Z]*)_TB$")

# DET 를 셀 때 뺄지 말지가 갈리는 감사 컬럼
AUDIT_COLUMNS = {"FRST_REGISTER_ID", "FRST_REGIST_DT",
                 "LAST_UPDUSR_ID", "LAST_UPDT_DT"}

# 테이블정의서 시트 이름. 앞부분만 맞으면 된다
SHEET_TABLES = "테이블목록"
SHEET_COLUMNS = "컬럼정의서"


# ======================================================================
# UI설계서 (pptx)
# ======================================================================
def _cell_text(tc):
    """표 한 칸의 글자를 잇는다. 전각 공백은 버린다."""
    return "".join(t.text or "" for t in tc.iter("{%s}t" % DRAWING_NS)
                   ).strip().replace("\u3000", "")


def _slide_tables(xml):
    """슬라이드 XML 하나에서 표를 모두 뽑는다. [[행][칸]] 목록."""
    root = _xml.fromstring(xml)
    return [[[_cell_text(tc) for tc in tr.findall("{%s}tc" % DRAWING_NS)]
             for tr in tb.findall("{%s}tr" % DRAWING_NS)]
            for tb in root.iter("{%s}tbl" % DRAWING_NS)]


def _norm(cells):
    """머리글 비교용. 공백과 각주 번호를 턴다."""
    return [c.replace(" ", "").replace("1", "") for c in cells]


def _is_screen_header(h):
    """화면 표지 표인가. UI ID · UI 명 이 있는 줄."""
    return len(h) >= 4 and h[0] == "UI ID" and h[2] == "UI 명"


def _is_attr_header(h):
    """화면속성 표인가. 그룹 · 화면속성 · Table 이 있는 줄."""
    n = _norm(h)
    return (len(n) >= 5 and n[1] == "그룹" and n[2] == "화면속성"
            and n[3] == "Table")


def _is_func_header(h):
    """처리상세 표인가. 기능 · 처리상세 가 있는 줄."""
    n = _norm(h)
    return len(n) >= 5 and n[1] == "기능" and n[2] == "처리상세"


def _slide_no(name):
    return int(re.search(r"(\d+)", name.split("/")[-1]).group(1))


def parse_ui_file(path, on_progress=None):
    """UI설계서 pptx 한 개를 읽는다.

    한 화면이 여러 슬라이드에 걸쳐 있을 수 있다. 화면 표지 표가 나온 뒤의
    속성·처리상세 표는 모두 그 화면 것으로 본다.

    돌려주는 것: ({화면ID: 화면}, [문제])
    """
    base = os.path.basename(path)
    screens, problems = {}, []

    with zipfile.ZipFile(path) as z:
        names = sorted(
            (n for n in z.namelist()
             if re.match(r"ppt/slides/slide\d+\.xml$", n)), key=_slide_no)
        total = len(names)
        current = None

        for i, name in enumerate(names, 1):
            no = _slide_no(name)
            try:
                tables = _slide_tables(z.read(name))
            except Exception as e:
                problems.append(f"{base} 슬라이드 {no}: 표를 읽지 못함 ({e})")
                continue

            for rows in tables:
                if rows and _is_screen_header(rows[0]):
                    sid = rows[0][1].strip()
                    if not SCREEN_ID.match(sid):
                        problems.append(f"{base} 슬라이드 {no}: "
                                        f"화면ID 형식이 아님 '{sid}'")
                    current = sid
                    s = screens.setdefault(sid, {
                        "id": sid, "name": rows[0][3].strip(), "file": base,
                        "slides": [], "attrs": [], "funcs": [], "tables": []})
                    s["slides"].append(no)

            if current is None:
                continue
            s = screens[current]
            for rows in tables:
                if not rows:
                    continue
                if _is_func_header(rows[0]):
                    for r in rows[1:]:
                        if any(r):
                            s["funcs"].append({
                                "seq": r[0] if len(r) > 0 else "",
                                "event": r[1] if len(r) > 1 else "",
                                "detail": r[2] if len(r) > 2 else ""})
                elif _is_attr_header(rows[0]):
                    for r in rows[1:]:
                        if len(r) < 5 or not any(r):
                            continue
                        s["attrs"].append({"name": r[2], "table": r[3].strip()})

            if on_progress and (i % 40 == 0 or i == total):
                on_progress(i, total)

    for s in screens.values():
        s["tables"] = sorted({a["table"] for a in s["attrs"]
                              if TABLE_NAME.match(a["table"])})
    return screens, problems


# ======================================================================
# 테이블정의서 (xlsx)
# ======================================================================
def _find_sheet(wb, part):
    """이름에 part 가 든 시트를 찾는다. '3.테이블목록' 처럼 번호가 붙어 있다."""
    for name in wb.sheetnames:
        if part in name.replace(" ", ""):
            return wb[name]
    raise KeyError(f"'{part}' 시트를 찾지 못했습니다. 있는 시트: "
                   f"{', '.join(wb.sheetnames)}")


def _find_header(ws, must, max_row=8):
    """머리글 줄을 찾는다. must 안의 이름이 모두 있는 줄이 머리글이다.

    돌려주는 것: (줄 번호, {열이름: 열 위치})
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row,
                                         values_only=True), 1):
        cells = [str(c).replace("\n", "").replace("*", "").strip()
                 if c is not None else "" for c in row]
        if all(any(m == c for c in cells) for m in must):
            return i, {c: j for j, c in enumerate(cells) if c}
    return None, None


def parse_table_file(path):
    """테이블정의서 xlsx 한 개를 읽는다.

    파일마다 물리명이 '테이블명' 칸에 있기도 하고 '엔터티명' 칸에 있기도 하다.
    두 칸 중 물리명 형식에 맞는 값이 더 많은 쪽을 물리명으로 본다.

    돌려주는 것: ({테이블명: 테이블}, [문제], 물리명이 있던 칸 이름)
    """
    from openpyxl import load_workbook

    base = os.path.basename(path)
    parts = base.split("_")
    system = parts[1] if len(parts) > 1 else os.path.splitext(base)[0]
    tables, problems = {}, []

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _find_sheet(wb, SHEET_TABLES)
        hr, h = _find_header(ws, ["테이블명", "엔터티명"])
        if hr is None:
            raise KeyError("테이블목록 시트에서 머리글을 찾지 못했습니다.")
        rows = [r for r in ws.iter_rows(min_row=hr + 1, values_only=True)
                if r and any(x is not None for x in r)]

        def cell(r, key):
            j = h.get(key)
            return "" if j is None or j >= len(r) else str(r[j] or "").strip()

        a = sum(1 for r in rows if TABLE_NAME.match(cell(r, "테이블명")))
        b = sum(1 for r in rows if TABLE_NAME.match(cell(r, "엔터티명")))
        phys, kor = ("테이블명", "엔터티명") if a >= b else ("엔터티명", "테이블명")

        for r in rows:
            name = cell(r, phys)
            if not name:
                continue
            if not TABLE_NAME.match(name):
                problems.append(f"{base}: 물리명 형식이 아님 '{name}'")
                continue
            tables[name] = {"name": name, "kor": cell(r, kor),
                            "system": system, "file": base, "columns": []}

        ws = _find_sheet(wb, SHEET_COLUMNS)
        hr, h = _find_header(ws, ["테이블명", "컬럼명"])
        if hr is None:
            raise KeyError("컬럼정의서 시트에서 머리글을 찾지 못했습니다.")
        seen = {}
        for r in ws.iter_rows(min_row=hr + 1, values_only=True):
            if not r:
                continue
            t = cell(r, "테이블명")
            c = cell(r, "컬럼명")
            if not t or not c or not TABLE_NAME.match(t):
                continue
            if t not in tables:
                problems.append(f"{base}: 컬럼정의서에만 있는 테이블 '{t}'")
                tables[t] = {"name": t, "kor": "", "system": system,
                             "file": base, "columns": []}
            s = seen.setdefault(t, set())
            if c not in s:          # 같은 컬럼명은 한 번만 센다
                s.add(c)
                tables[t]["columns"].append(c)
    finally:
        wb.close()

    return tables, problems, phys


# ======================================================================
# 논리파일 그룹
# ======================================================================
def group_key(name):
    """테이블명에서 논리파일 그룹 키를 만든다.

    기준코드가 같은 테이블을 한 묶음으로 본다. 예)
      COM_CM001A_TB, COM_CM001B_TB → COM_CM_001
    형식에 안 맞으면 테이블명 자체를 키로 쓴다.
    """
    m = GROUP_KEY.match(name)
    return "_".join(m.group(1, 2, 3)) if m else name


def build_groups(tables):
    """테이블을 그룹으로 묶고 RET·DET 후보를 센다.

    RET 는 그룹에 속한 테이블 수, DET 는 컬럼명 합집합의 크기다.
    이 값은 후보일 뿐이다. 한 논리파일로 볼지는 감정인이 판단한다.

    돌려주는 것: ({그룹키: 그룹}, {테이블명: 그룹키})
    """
    bucket = {}
    for name in tables:
        bucket.setdefault(group_key(name), []).append(name)

    groups = {}
    for gk, names in bucket.items():
        cols = set()
        for n in names:
            cols |= set(tables[n]["columns"])
        groups[gk] = {"key": gk, "tables": sorted(names),
                      "ret": len(names), "det": len(cols),
                      "det_no_audit": len(cols - AUDIT_COLUMNS)}
    return groups, {n: gk for gk, names in bucket.items() for n in names}


def screen_events(screen, limit=14):
    """화면의 단위프로세스명을 모은다.

    UI설계서 처리상세 표의 '기능' 칸이 단위프로세스에 해당한다. 괄호 안 설명은
    떼고, 같은 이름은 한 번만 남긴다. FP 산정 시트의 단위프로세스명 칸과
    맞춰 볼 때 쓴다.
    """
    out = []
    for f in screen.get("funcs", []):
        name = re.sub(r"\s*\(.*", "", str(f.get("event") or "")).strip()
        if name and name not in out:
            out.append(name)
    return out[:limit]


# ======================================================================
# 화면-테이블 매핑표
#   FTR 판정의 근거표다. 저장하는 JSON 에는 넣지 않는다. screens 와 tables 로
#   언제든 다시 만들 수 있고, 같은 값을 두 곳에 두면 어긋나기 때문이다.
# ======================================================================
def build_mapping(screens, tables, groups, table_group):
    """화면과 테이블을 짝지어 두 가지 표를 만든다.

    pairs   한 줄이 (화면, 테이블) 한 쌍. 그 화면 속성 중 이 테이블을 가리키는
            행이 몇 개인지(attrs)까지 센다. DET 를 따질 때 쓴다.
    screens 화면 한 줄. FTR 후보와 DET 상한을 담는다.

    FTR 후보는 물리 테이블 수가 아니라 논리파일 그룹 수다. 기준코드가 같은
    테이블을 여러 개 쓰는 화면은 물리 테이블 수보다 FTR 후보가 작다.
    한 논리파일로 볼지는 감정인이 판단한다.
    """
    pairs, summary = [], []
    for s in sorted(screens.values(), key=lambda x: x["id"]):
        events = " / ".join(screen_events(s))
        per_table = {}
        for a in s["attrs"]:
            if a["table"] in s["tables"]:
                per_table[a["table"]] = per_table.get(a["table"], 0) + 1

        gset = set()
        for name in s["tables"]:
            t = tables.get(name)
            gk = table_group.get(name, "")
            g = groups.get(gk, {})
            if gk:
                gset.add(gk)
            pairs.append({
                "screen": s["id"], "screen_name": s["name"],
                "screen_file": s["file"], "events": events,
                "table": name, "kor": t["kor"] if t else "",
                "system": t["system"] if t else "",
                "known": 1 if t else 0,
                "group": gk, "group_n": len(g.get("tables", [])),
                "group_det": g.get("det"),
                "cols": len(t["columns"]) if t else None,
                "attrs": per_table.get(name, 0)})

        if not s["tables"]:
            pairs.append({
                "screen": s["id"], "screen_name": s["name"],
                "screen_file": s["file"], "events": events,
                "table": "", "kor": "", "system": "", "known": 0,
                "group": "", "group_n": 0, "group_det": None,
                "cols": None, "attrs": 0})

        summary.append({
            "screen": s["id"], "name": s["name"], "file": s["file"],
            "events": events,
            "tables": len(s["tables"]),
            "ftr": len(gset),                       # FTR 후보 = 논리파일 그룹 수
            "det_max": len(s["attrs"]),             # DET 상한 = 화면 속성 행 수
            "funcs": len(s["funcs"]),
            "missing": sum(1 for n in s["tables"] if n not in tables)})
    return pairs, summary


MAPPING_COLUMNS = [
    ("screen", "화면ID"), ("screen_name", "화면명"),
    ("events", "단위프로세스명"), ("screen_file", "출처파일"),
    ("table", "테이블명"), ("kor", "한글명"), ("system", "업무"),
    ("known", "정의서존재"), ("group", "논리파일그룹"),
    ("group_n", "그룹내표수"), ("group_det", "그룹DET"),
    ("cols", "표컬럼수"), ("attrs", "이표참조속성수"),
]


def mapping_csv(pairs):
    """매핑표를 CSV 글자열로 만든다. 엑셀에서 열 수 있도록 BOM 을 붙인다."""
    import csv
    import io
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\r\n")
    w.writerow([label for _, label in MAPPING_COLUMNS])
    for r in pairs:
        w.writerow(["" if r.get(k) is None else r.get(k)
                    for k, _ in MAPPING_COLUMNS])
    return "\ufeff" + buf.getvalue()


# ======================================================================
# 폴더 단위로 읽기
# ======================================================================
def _read_dir(pattern, folder, kind, reader, cache, on_file, merge):
    files = sorted(glob.glob(os.path.join(folder, pattern)))
    if not files:
        raise FileNotFoundError(f"{folder} 에서 {pattern} 을 찾지 못했습니다.")
    out, problems, report = {}, [], []
    for i, path in enumerate(files):
        t0 = time.time()
        hit = cache.get(kind, path) if cache else None
        if hit is not None:
            data, probs, note = hit["data"], hit["problems"], hit.get("note", "")
            cached = True
        else:
            data, probs, note = reader(path, i)
            if cache:
                cache.put(kind, path, {"data": data, "problems": probs,
                                       "note": note})
            cached = False
        merge(out, data)
        problems += probs
        report.append({"name": os.path.basename(path), "kind": kind,
                       "count": len(data), "cached": cached, "note": note,
                       "ms": int((time.time() - t0) * 1000)})
        if on_file:
            on_file(report[-1], i, len(files))
    return out, problems, report


def load_screens(folder, cache=None, on_file=None, on_progress=None):
    """폴더 안의 UI설계서를 모두 읽어 합친다."""
    def reader(path, i):
        screens, probs = parse_ui_file(
            path, (lambda c, t: on_progress(i, c, t)) if on_progress else None)
        return screens, probs, ""

    def merge(out, data):
        for k, v in data.items():
            if k in out:
                out[k]["attrs"] += v["attrs"]
                out[k]["funcs"] += v["funcs"]
                out[k]["slides"] += v["slides"]
                out[k]["tables"] = sorted(set(out[k]["tables"]) | set(v["tables"]))
            else:
                out[k] = v

    return _read_dir(UI_GLOB, folder, "화면", reader, cache, on_file, merge)


def load_tables(folder, cache=None, on_file=None):
    """폴더 안의 테이블정의서를 모두 읽어 합친다."""
    def reader(path, i):
        tables, probs, phys = parse_table_file(path)
        return tables, probs, f"물리명은 '{phys}' 칸"

    def merge(out, data):
        for k, v in data.items():
            if k in out:
                have = set(out[k]["columns"])
                out[k]["columns"] += [c for c in v["columns"] if c not in have]
            else:
                out[k] = v

    return _read_dir(TBL_GLOB, folder, "테이블", reader, cache, on_file, merge)



# ======================================================================
# 저장 · 적재
#   다른 모듈은 산출물(pptx·xlsx)을 다시 읽지 않는다. 이 JSON 만 읽는다.
# ======================================================================
FORMAT = "fp-checker/doc-parser"
FORMAT_VERSION = 1

# 저장 파일 이름. 읽은 시각을 붙여 남긴다. 앞에서부터 읽으면 시간 순서가 된다
NAME_PREFIX = "doc_parser"
NAME_TIME = "%Y%m%d_%H%M"
NAME_GLOB = NAME_PREFIX + "_*.json"

# 가장 최근 결과를 늘 같은 이름으로도 남긴다. 다른 모듈은 이 이름만 알면 된다.
# 시각이 붙은 파일은 이력으로 남는다.
LATEST_NAME = NAME_PREFIX + ".json"


def make_name(when=None):
    """저장 파일 이름을 만든다. 예) doc_parser_20260903_1432.json"""
    when = when or datetime.datetime.now()
    return f"{NAME_PREFIX}_{when.strftime(NAME_TIME)}.json"


def find_all(folder):
    """폴더 안의 저장 파일을 오래된 것부터 늘어놓는다."""
    return sorted(glob.glob(os.path.join(os.fspath(folder), NAME_GLOB)))


def update_latest(folder, src):
    """방금 저장한 것을 doc_parser.json 으로도 복사한다.

    쓰다 만 파일이 남지 않도록 임시 이름으로 복사한 뒤 바꿔치운다.
    """
    dst = os.path.join(os.fspath(folder), LATEST_NAME)
    tmp = dst + ".tmp"
    shutil.copyfile(os.fspath(src), tmp)
    os.replace(tmp, dst)
    return dst


def latest(folder):
    """가장 최근 결과의 경로. 없으면 None.

    doc_parser.json 이 있으면 그것을 쓴다. 없으면 시각이 붙은 파일 중 최신을
    쓴다. 복사가 실패했거나 그 파일만 지운 경우를 대비한 것이다.

    다른 모듈은 이것으로 파일을 찾는다.
      path = doc_parser.latest(common.out_dir())
      data = doc_parser.load(path)
    """
    fixed = os.path.join(os.fspath(folder), LATEST_NAME)
    if os.path.exists(fixed):
        return fixed
    files = find_all(folder)
    return files[-1] if files else None


def prune(folder, keep=0):
    """오래된 저장 파일을 지운다. keep 이 0 이면 아무것도 지우지 않는다."""
    if keep <= 0:
        return []
    files = find_all(folder)
    gone = []
    for path in files[:-keep]:
        try:
            os.remove(path)
            gone.append(path)
        except OSError:
            pass
    return gone


def dump(screens, tables, groups, table_group, problems=None, source=None):
    """다른 모듈에 넘길 한 벌을 만든다.

    {
      "format": "fp-checker/doc-parser", "version": 1,
      "created": "2026-09-03T09:20:11",
      "source": {"ui_dir","tbl_dir","files":[{name,kind,count,ms,cached,note}]},
      "counts": {"screens","tables","columns","groups","problems"},
      "screens":     {화면ID: {id,name,file,slides,attrs,funcs,tables}},
      "tables":      {테이블명: {name,kor,system,file,columns}},
      "groups":      {그룹키: {key,tables,ret,det,det_no_audit}},
      "table_group": {테이블명: 그룹키},
      "problems":    ["..."]
    }

    screens 의 tables 는 화면속성표에 적힌 물리 테이블이다. 그 테이블이
    테이블정의서에 실제로 있는지는 확인하지 않는다. 받는 쪽에서 맞춰 보라.
    """
    return {
        "format": FORMAT,
        "version": FORMAT_VERSION,
        "created": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": source or {},
        "counts": {
            "screens": len(screens), "tables": len(tables),
            "columns": sum(len(t["columns"]) for t in tables.values()),
            "groups": len(groups), "problems": len(problems or [])},
        "screens": screens,
        "tables": tables,
        "groups": groups,
        "table_group": table_group,
        "problems": list(problems or []),
    }


def save(path, data, pretty=False):
    """JSON 으로 적는다. 쓰다 만 파일이 남지 않도록 임시 파일에 쓰고 바꿔치운다."""
    path = os.fspath(path)
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False,
                  indent=2 if pretty else None)
    os.replace(tmp, path)
    return path


def load(path):
    """저장해 둔 JSON 을 읽는다. 형식이 다르면 바로 알려 준다."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if data.get("format") != FORMAT:
        raise ValueError(f"{os.path.basename(path)} 는 doc_parser 가 만든 "
                         f"파일이 아닙니다. (format={data.get('format')!r})")
    if data.get("version") != FORMAT_VERSION:
        raise ValueError(f"형식 판이 다릅니다. 이 프로그램은 {FORMAT_VERSION} 판을 "
                         f"읽습니다. 파일은 {data.get('version')} 판입니다. "
                         f"산출물을 다시 읽어 저장하십시오.")
    for key in ("screens", "tables", "groups", "table_group"):
        if not isinstance(data.get(key), dict):
            raise ValueError(f"'{key}' 가 없거나 형식이 맞지 않습니다.")
    return data


def read_and_save(ui_dir, tbl_dir, out_path, cache=None, on_file=None,
                  on_progress=None, pretty=False):
    """두 폴더를 읽어 JSON 으로 저장하고, 만든 한 벌을 돌려준다."""
    screens, sp, ui_rep = load_screens(ui_dir, cache, on_file, on_progress)
    tables, tp, tbl_rep = load_tables(tbl_dir, cache, on_file)
    groups, table_group = build_groups(tables)
    data = dump(screens, tables, groups, table_group, sp + tp,
                {"ui_dir": ui_dir, "tbl_dir": tbl_dir,
                 "files": ui_rep + tbl_rep})
    save(out_path, data, pretty)
    return data


# ======================================================================
# 단독 실행
# ======================================================================
def _main(argv):
    args = [a for a in argv[1:] if not a.startswith("--")]
    pretty = "--pretty" in argv
    if len(args) < 2:
        print(__doc__)
        return 1
    ui_dir, tbl_dir = args[0], args[1]
    out_path = args[2] if len(args) > 2 else None

    print(f"XML 파서: {XML_BACKEND}")
    screens, sp, _ = load_screens(ui_dir)
    tables, tp, _ = load_tables(tbl_dir)
    groups, table_group = build_groups(tables)
    problems = sp + tp

    cols = sum(len(t["columns"]) for t in tables.values())
    print(f"화면 {len(screens)}개 · 테이블 {len(tables)}개 · "
          f"컬럼 {cols}개 · 논리파일 그룹 {len(groups)}개")
    known = set(tables)
    used = {t for s in screens.values() for t in s["tables"]}
    print(f"화면이 쓰는 테이블 {len(used & known)}개 · "
          f"정의서에 없는 것 {len(used - known)}개 · "
          f"화면에 안 나오는 것 {len(known - used)}개")
    for p in problems[:20]:
        print("  ·", p)
    if len(problems) > 20:
        print(f"  … 문제 {len(problems)}건 중 20건만 보였습니다.")

    pairs, summary = build_mapping(screens, tables, groups, table_group)
    over = sum(1 for r in summary if r["ftr"] < r["tables"])
    print(f"매핑 {len(pairs)}쌍 · 물리 표 수보다 FTR 후보가 작은 화면 {over}개")

    if out_path:
        # .json 으로 끝나지 않으면 폴더로 본다. 아직 없는 폴더도 마찬가지다
        if not out_path.lower().endswith(".json"):
            out_path = os.path.join(out_path, make_name())
        data = dump(screens, tables, groups, table_group, problems,
                    {"ui_dir": ui_dir, "tbl_dir": tbl_dir})
        save(out_path, data, pretty)
        size = os.path.getsize(out_path) / 1024 / 1024
        print(f"저장: {out_path} ({size:.1f} MB)")
        fixed = update_latest(os.path.dirname(os.path.abspath(out_path)), out_path)
        print(f"최신본: {fixed}")
    return 0


if __name__ == "__main__":
    sys.exit(_main(sys.argv))